# -*- coding: utf-8 -*-

#-----------------------------------------------------------------------------
# Controladores provisionales utilizados solo para probar las vistas del modulo de SMyDP
#
# - Samuel Arleo <saar1312@gmail.com>
# 
# - Convenciones:
# * Funciones "privadas" utilizadas por los controladores y decoradores tienen el
# prefijo "__"
# * Controladores no poseen prefijos
#
#-----------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import unicodedata
import calendar
import datetime
from sustancias_libreria import *


# Verifica si el usuario que intenta acceder al controlador tiene alguno de los
# roles necesarios
def __check_role():

    roles_permitidos = ['WEBMASTER', 'DIRECTOR', 'ASISTENTE DEL DIRECTOR', 
                        'JEFE DE LABORATORIO', 'JEFE DE SECCIÓN', 'TÉCNICO', 
                        'GESTOR DE SMyDP', 'PERSONAL INTERNO']
    return True in map(lambda x: auth.has_membership(x), roles_permitidos)

def __check_role_report():

    roles_permitidos = ['WEBMASTER', 'DIRECTOR', 'ASISTENTE DEL DIRECTOR', 
                                'GESTOR DE SMyDP']
    return True in map(lambda x: auth.has_membership(x), roles_permitidos)

# Determina si el id de la dependencia es valido. Retorna False si el id no existe
# o es de un tipo incorrecto
def __is_valid_id(id_, tabla):
    try:
        int(id_)
    except:
        return False
    # Si el id recibido tiene el tipo correcto pero no existe en la base de datos
    if not db(tabla.id == int(id_)).select():
        return False

    return True

# Determina si una variable "booleana" pasada como parametro con GET es realmente
# 'True' o 'False' (request.vars almacena todo como strings)
def __is_bool(bool_var):
    if not bool_var in ['True', 'False']:
        return False
    else:
        return True

# Dado el nombre de una dependencia, retorna el id de esta si la encuentra o
# None si no lo hace
def __find_dep_id(nombre):

    dep_id = db(db.dependencias.nombre == nombre).select()[0].id    
    return dep_id

# Dado el id de un espacio fisico, retorna las sustancias que componen el inventario
# de ese espacio.
def __get_inventario_espacio(espacio_id=None):
    inventario = []
    inventario = list(db((db.t_Inventario.sustancia == db.t_Sustancia.id) &
                         (db.t_Inventario.f_medida == db.t_Unidad_de_medida.id) & 
                         (db.t_Inventario.espacio == espacio_id)).select())

    return inventario

# Dado el id de un espacio fisico, retorna los desechos peligrosos que componen el inventario
# de ese espacio. Si ningun id es indicado, pero si el de una dependencia, busca
# todos los espacios fisicos que pertenecen a esta, agrega los inventarios y retorna
# la lista
def __get_inventario_desechos(espacio_id=None, dep_id=None):
    inventario = []
    if espacio_id:
        inventario = list(db((db.t_inventario_desechos.grupo == db.t_categoria_desechos.id) &
                         (db.t_inventario_desechos.unidad_medida == db.t_Unidad_de_medida.id) & 
                         (db.t_inventario_desechos.espacio_fisico == espacio_id)).select())
    
    return inventario

# Retorna las hojas o dependencias que no tienen hijos (posiblemente secciones) y
# que estan por debajo de la dependencia dada.
# "jerarquia" tiene la forma: 
#       {'dependencia1': [dep_hija1,
#                        .
#                        .
#                         dep_hijan]
#        'dependencia2': [dep_hija1,
#                         .
#                         .
#                         dep_hijam]
#     }
# Si una dependencia no tiene otras adscritas, entonces no aparece en "jerarquia"
def __get_leaves(dep_id, jerarquia):

    if not dep_id in jerarquia:
        return [dep_id]
    else:
        l = []
        for d in jerarquia[dep_id]:
            l = l + __get_leaves(d, jerarquia) 
        return l


# Dada una lista de ids de dependencias que no poseen otras adscritas a ellas,
# retorna los ids de espacios fisicos en la base de datos que tienen a estas 
# dependencias como secciones
def __filtrar_espacios(hojas):

    espacios = []
    for dep_id in hojas:
        nuevos_espacios = [esp.id for esp in db(db.espacios_fisicos.dependencia == dep_id).select()]
        if nuevos_espacios:
            espacios = espacios + nuevos_espacios
    return espacios

# Dado el id de una dependencia, retorna una lista con los ids de todos los
# espacios fisicos que pertenecen a esta. Si el id es de la ULAB, retorna
# todos los espacios fisicos
def __get_espacios(dep_id):
    espacios = []

    secciones = []
    dependencias = db(db.dependencias.id > 0).select()
    
    # Creando lista de adyacencias
    lista_adyacencias = {dep.id: dep.unidad_de_adscripcion for dep in dependencias}

    # Representando la jerarquia con la forma {'dependencia': [dep_hija1, dep_hija2]}
    jerarquia = {}

    for hijo, padre in lista_adyacencias.iteritems():
        # Si el padre es None, es porque se trata de la unidad de laboratorios
        # que no tiene padre (nivel mas alto de la jerarquia)
        if padre is not None:
            if padre in jerarquia:
                jerarquia[padre].append(hijo)
            else:
                jerarquia[padre] = [hijo]

    hojas = __get_leaves(int(dep_id), jerarquia)

    espacios = __filtrar_espacios(hojas)

    return espacios


# Transforma una cantidad en la unidad de medida indicada en nueva_unidad
def __transformar_cantidad(cantidad, unidad, nueva_unidad):
    cantidad = float(cantidad)
    if nueva_unidad == unidad:
        return cantidad
    elif unidad in ["Kilogramos", "Litros"]:
        return cantidad * 1000
    elif unidad in ["Mililitros", "Gramos"]:
        return cantidad / 1000

# Permite sumar dos cantidades de sustancia de acuerdo a la unidad en la que
# se esta mostrando la cantidad de sustancia en el inventario. *!* Se asume que
# no se agregaran sustancias a los inventarios en unidades diferentes a las normales
# Si es un liquido, se puede pasar de litros a mililitros, pero no se espera
# que se ingrese en algun inventario en gramos. Siempre se pasa a la unidad de 
# medida que ya estaba
def __sumar_cantidad(nueva_cantidad, cantidad_actual, nueva_unidad, unidad):

    if nueva_unidad == unidad:
        return float(nueva_cantidad) + float(cantidad_actual)
    # Si no son iguales y ademas la nueva sustancia esta en Litros o Kilos
    elif nueva_unidad in ["Kilogramos", "Litros"]:
        return float(nueva_cantidad)*1000 + float(cantidad_actual)
    # Si no son iguales y ademas la nueva sustancia esta en Mililitros o gramos
    elif nueva_unidad in ["Mililitros", "Gramos"]:
        return float(nueva_cantidad)/1000 + float(cantidad_actual)


# Agrega los inventarios de los espacios en la lista "espacios"
def __sumar_inventarios(espacios):

    inventario_total = {}
    for esp_id in espacios:
        # Recorriendo las entradas en el inventario que pertenecen al espacio "esp"
        for row in db((db.t_Inventario.sustancia == db.t_Sustancia.id) &
                      (db.t_Inventario.f_medida == db.t_Unidad_de_medida.id) & 
                      (db.t_Inventario.espacio == esp_id)).select():

            sust = row['t_Sustancia']
            inv = row['t_Inventario']
            unid = row['t_Unidad_de_medida']

            sustancia_id = sust.id

            # Se agrega la sustancia al inventario final si esta no estaba ya
            if not sustancia_id in inventario_total:
                
                inventario_total[sustancia_id] = {
                                        'f_nombre': sust.f_nombre,
                                        'f_cas': sust.f_cas,
                                        'f_pureza': sust.f_pureza,
                                        'f_estado': sust.f_estado,
                                        'f_existencia':inv.f_existencia,
                                        'f_uso_interno': inv.f_uso_interno,
                                        'f_unidad': unid.f_nombre
                                                 }
            # Si ya estaba, se suma la cantidad en existencia y de uso interno
            # de la sustancia con id sustancia_id
            else:
                # Inventario actual de la sustancia sustancia_id
                s = inventario_total[sustancia_id]

                # Cantidades existentes por ahora en el inventario general
                existencia = s['f_existencia']
                uso_interno = s['f_uso_interno']

                # Unidad en que se mostrara el inventario general de la sustancia
                unidad = s['f_unidad']

                # Nuevas cantidades que hay que sumar al inventario general
                nueva_exist = inv.f_existencia
                nuevo_uso_interno = inv.f_uso_interno
                nueva_unidad = unid.f_nombre                
                s['f_existencia'] = __sumar_cantidad(nueva_exist,
                                                    existencia,
                                                    nueva_unidad,
                                                    unidad)
                s['f_uso_interno'] = __sumar_cantidad(nuevo_uso_interno,
                                                     uso_interno,
                                                     nueva_unidad,
                                                     unidad)
                    
    return inventario_total


# consulta los espacios que tienen cierta sustancia
def __consultar_espacios(espacios, sustancia):

    inventario_total = {}
    i = 0
    for esp_id in espacios:
        # Recorriendo las entradas en el inventario que pertenecen al espacio "esp"
        for row in db((db.t_Inventario.sustancia == db.t_Sustancia.id) &
                      (db.t_Inventario.f_medida == db.t_Unidad_de_medida.id) & 
                      (db.t_Inventario.espacio == esp_id)).select():

            esp_actual = db(db.espacios_fisicos.id == esp_id).select().first()

            sust = row['t_Sustancia']
            inv = row['t_Inventario']
            unid = row['t_Unidad_de_medida']

            # Se agrega la sustancia al inventario final
            if sust.f_nombre == sustancia:
                
                i += 1
                inventario_total[int(i)] = {
                                    'f_inv': inv.id,
                                    'f_espacio': esp_actual.codigo,
                                    'f_cas': sust.f_cas,
                                    'f_pureza': sust.f_pureza,
                                    'f_estado': sust.f_estado,
                                    'f_existencia':inv.f_existencia,
                                    'f_uso_interno': inv.f_uso_interno,
                                    'f_unidad': unid.f_nombre
                                             }
                    
    return inventario_total


# Dado el id de una dependencia, retorna una lista con el agregado de las sutancias
# que existen en los espacios fisicos que pertenecen a esta. 
def __get_inventario_dep(dep_id):

    inventario = {}

    # Obteniendo lista de espacios bajo la dependencia con id dep_id
    espacios = __get_espacios(dep_id)

    # Agrega los inventarios de los espacios en la lista "espacios"
    inventario = __sumar_inventarios(espacios)

    return inventario


# Dado el id de una dependencia, retorna una lista con el agregado de las sutancias
# que existen en los espacios fisicos que pertenecen a esta. 
def __solicitar_inventario_dep(dep_id, sustancia):

    inventario = {}

    # Obteniendo lista de espacios bajo la dependencia con id dep_id
    espacios = __get_espacios(dep_id)

    # Agrega los inventarios de los espacios en la lista "espacios"
    inventario = __consultar_espacios(espacios, sustancia)

    return inventario


# Registra una nueva sustancia en el espacio fisico indicado. Si la sustancia ya
# existe en el inventario, genera un mensaje con flash y no anade de nuevo la
# sustancia. 
def __agregar_sustancia(espacio, sustancia_id, total, uso_interno, unidad_id):

    # Si ya existe la sustancia en el inventario
    if db((db.t_Inventario.espacio == espacio.id) & 
          (db.t_Inventario.sustancia == sustancia_id)).select():
        sust = db(db.t_Sustancia.id == sustancia_id).select()[0]

        response.flash = "La sustancia \"{0}\" ya ha sido ingresada anteriormente \
                          al espacio \"{1}\".".format(sust.f_nombre, espacio.codigo)
        return False
    elif(int(total)-int(uso_interno)<0):
        response.flash = "No puede haber mas sustancia en uso que en total"
        return False 

    # Si no, se agrega al inventario del espacio fisico la nueva sustancia
    else:
        cantidad = float(total)
        inv_id = db.t_Inventario.insert(f_existencia=cantidad, 
                                f_uso_interno=float(uso_interno),
                                f_medida=unidad_id,
                                espacio=espacio.id,
                                sustancia=sustancia_id)

        concepto = 'Ingreso'
        tipo_ing = 'Ingreso inicial'

        # Agregando la primera entrada de la sustancia en la bitacora
        db.t_Balance.insert(
                                f_cantidad=cantidad,
                                f_cantidad_total=cantidad,
                                f_concepto=concepto,
                                f_tipo_ingreso=tipo_ing,
                                f_medida=unidad_id,
                                f_inventario=inv_id,
                                f_sustancia=sustancia_id,
                                f_fechaUso=datetime.date.today())

    return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

# Dado el id de una depencia y conociendo si es un espacio fisico o una dependencia
# comun, determina si el usuario tiene privilegios suficientes para obtener informacion
# de esta
def __acceso_permitido(user, dep_id, es_espacio):
    """
    Args:
        * user_id (str): id del usuario en la tabla t_Personal (diferente de auth.user.id)
        * dep_id (str): id de la dependencia a la cual pertenece el recurso que se 
            desea acceder
        * es_espacio (str): 'True' si el usuario viene de seleccionar un espacio 
            fisico
    """
    # Valor a retornar que determina si el usuario tiene o no acceso al recurso
    permitido = False

    # dep_actual es un apuntador que permitira recorrer la jerarquia de dependencias
    # desde dep_id hasta usuario_dep. Si dep_actual no encuentra usuario_dep 
    # entonces se esta tratando de acceder a una dependencia sin permisos suficientes
    dep_actual = dep_id

    # Si el usuario es tecnico se busca en la tabla de es_encargado si el usuario 
    # es encargado del espacio con id dep_id
    if auth.has_membership("TÉCNICO"):
        encargado = db(db.es_encargado.espacio_fisico == dep_id).select().first()
        if encargado:
            permitido = encargado.tecnico == user.id

    else:
        # Dependencia a la que pertenece el usuario o que tiene a cargo
        usuario_dep = user.f_dependencia

        # Buscando todas las dependencias para conocer la lista de adyacencias con
        # la jerarquia de la ULAB
        dependencias = db(db.dependencias.id > 0).select(
                                db.dependencias.nombre,
                                db.dependencias.id,
                                db.dependencias.unidad_de_adscripcion)
        

        # Creando lista de adyacencias
        lista_adyacencias = {dep.id: dep.unidad_de_adscripcion for dep in dependencias}

        # Buscando el id de la direccion para saber si ya se llego a la raiz
        direccion_id = __find_dep_id('DIRECCIÓN')

        # Si dep_id es un espacio fisico, se sube un nivel en la jerarquia (hasta
        # las secciones) ya que los espacios fisicos no aparecen en la lista de 
        # adyacencias pero si las secciones a las que pertenecen
        if es_espacio == "True":
            dep_actual = db(db.espacios_fisicos.id == dep_id).select().first().dependencia

        while dep_actual is not None:

            # Si en el camino hacia la raiz se encontro la dependencia a la que
            # pertenece el usuario, entonces si hay privilegios suficientes
            if dep_actual == usuario_dep:
                permitido = True
                break
            # Si ya se llego a la raiz, terminar el while
            if dep_actual == direccion_id:
                break
            else:
                dep_actual = lista_adyacencias[dep_actual] 

    return permitido

# Retorna un string con la descripcion de un registro de la bitacora de acuerdo 
# a si es un ingreso (sompra, suministro almacen u otorgado por otra seccion) 
# o un egreso (docencia, invenstigacion o extension)
def __get_descripcion(registro):
    descripcion = ""

    if registro.f_concepto[0] == "Ingreso":
        # Si es un ingreso por compra, se muestra el 
        # Compra a "Proveedor" según Factura No. "No. Factura" de fecha "Fecha de compra"
        if registro.f_tipo_ingreso[0] == "Compra":
            compra = db(db.t_Compra.id == registro.f_compra).select()[0]
            
            # Datos de la compra
            proveedor = compra.f_institucion
            nro_factura = compra.f_nro_factura
            fecha_compra = compra.f_fecha

            fecha = fecha_compra

            descripcion = "Compra a \"{0}\" según Factura No. \"{1}\" con fecha"\
                         " \"{2}\"".format(proveedor, nro_factura, fecha)

        # Si es un ingreso por almacen
        # Suministro por el almacén del Laboratorio "X" 
        elif registro.f_tipo_ingreso[0] == "Almacén":
            almacen = db(db.espacios_fisicos.id == registro.f_almacen).select()[0]
            dep_id = almacen.dependencia
            dep = db(db.dependencias.id == dep_id).select()[0]

            # Asumiendo que siempre habra un laboratorio sobre la seccion a la que
            descripcion = "Suministrado por el almacén de la dependencia "\
                          "\"{0}\"".format(dep.nombre)

        elif registro.f_tipo_ingreso[0] == "Solicitud":
            # Respuesta a la solicitud en la que se otorgo la sustancia
            respuesta = db(db.t_Respuesta.id == registro.f_respuesta_solicitud
                          ).select()[0]

            # Espacio desde el que se acepto proveer la sustancia
            espacio = db(db.espacios_fisicos.id == respuesta.f_espacio).select()[0]

            # Seccion a la que pertenece ese espacio
            seccion = db(db.dependencias.id == espacio.dependencia).select()[0]

            # Laboratorio al que pertenece esa seccion
            lab = db(db.dependencias.id == seccion.unidad_de_adscripcion).select()[0]

            descripcion = "Otorgado por la Sección \"{0}\" del \"{1}\" "\
                          "en calidad de \"{2}\"".format(seccion.nombre,
                          lab.nombre, respuesta.f_calidad[0])


        elif registro.f_tipo_ingreso[0] == "Prestamo":
            descripcion = "Ingreso por prestamo "

        elif registro.f_tipo_ingreso[0] == "Cesión":
            descripcion = "Ingreso por Cesión "


        elif registro.f_tipo_ingreso[0] == "Ingreso inicial":
            descripcion = "Ingreso inicial de la sustancia al inventario"

    else:
        # Si es un consumo por Docencia
        if registro.f_tipo_egreso[0] == "Docencia":
            servicio = db(db.servicios.id == registro.f_servicio).select()[0] 

            nombre = servicio.nombre

            descripcion = "Ejecución de la práctica \"{0}\"".format(nombre)
        elif registro.f_tipo_egreso[0] == "Investigación":
            servicio = db(db.servicios.id == registro.f_servicio).select()[0] 

            nombre = servicio.nombre

            descripcion = "Ejecución del proyecto de investigación \"{0}\"".format(nombre)
            
        elif registro.f_tipo_egreso[0] == "Extensión":
            servicio = db(db.servicios.id == registro.f_servicio).select()[0] 

            nombre = servicio.nombre

            descripcion = "Ejecución del servicio \"{0}\"".format(nombre)


        elif registro.f_tipo_egreso[0] == "Préstamo":
           
            descripcion = "Egreso por prestamo "

        elif registro.f_tipo_egreso[0] == "Cesión":
        
            descripcion = "Egreso por Cesión "
        # Cuando es un egreso en respuesta a una solicitud
        elif registro.f_tipo_egreso[0] == "Solicitud":
       
            
            # Respuesta a la solicitud en la que se solicito la sustancia
            respuesta = db(db.t_Respuesta.id == registro.f_respuesta_solicitud
                          ).select()[0]

            # Solicitud que hizo que por aceptarla se sacara material
            solicitud = db(db.t_Solicitud_smydp.id == respuesta.f_solicitud
                          ).select()[0]

            # Espacio desde el que se solicito la sustancia
            espacio = db(db.espacios_fisicos.id == solicitud.f_espacio).select()[0]

            # Seccion a la que pertenece ese espacio
            seccion = db(db.dependencias.id == espacio.dependencia).select()[0]

            # Laboratorio al que pertenece esa seccion
            lab = db(db.dependencias.id == seccion.unidad_de_adscripcion).select()[0]

            descripcion = "Otorgado a la Sección \"{0}\" del \"{1}\" "\
                          "en calidad de \"{2}\"".format(seccion.nombre,
                          lab.nombre, respuesta.f_calidad[0])
        

    return descripcion

# Agrega un nuevo registro a la bitacora de una sustancia
def __agregar_registro(concepto):

    cantidad = float(request.vars.cantidad)

    # Operaciones comunes a todos los casos: actualizacion del inventario

    # ID de la unidad en la que el usuario registro la cantidad ingresada
    unidad_id = request.vars.unidad

    # Inventario al cual pertenece la bitacora consultada
    inv = db(db.t_Inventario.id == request.get_vars.inv).select()[0]

    # Unidad indicada por el usuario
    unidad = db(db.t_Unidad_de_medida.id == unidad_id
                   ).select()[0].f_nombre

    # Unidad de medida en la que se encuentra el inventario de la sustancia
    unidad_inventario = db(db.t_Unidad_de_medida.id == inv.f_medida
                          ).select()[0].f_nombre

    # Transformando las cantidades de acuerdo a la unidad utilizada en
    # el inventario de la sustancia
    cantidad = __transformar_cantidad(cantidad, unidad, unidad_inventario)


    inventario_id = int(request.vars.inv)

    bitacora = db((db.t_Balance.f_inventario == inventario_id) &
                  (db.t_Balance.created_by == db.auth_user.id) &
                  (db.auth_user.id == db.t_Personal.f_usuario) &
                  (db.t_Balance.f_medida == db.t_Unidad_de_medida.id)).select(orderby=~db.t_Balance.f_fechaUso)
    auxIng=0
    auxEgr=0

            
        
        
    # Cantidades total y de uso interno antes del ingreso o consumo
    
    total_viejo = inv.f_existencia
    uso_interno_viejo = inv.f_uso_interno
 
    # Nueva cantidad total y nueva cantidad para uso interno


    if concepto == 'Ingreso':
        total_nuevo = total_viejo + cantidad
        uso_interno_nuevo = uso_interno_viejo + cantidad
        
        inv.update_record(
                f_existencia=total_nuevo,
                f_uso_interno=uso_interno_nuevo)

        tipo_ing = request.vars.tipo_ingreso
        # Actualizando cantidad total con la nueva 
        if tipo_ing == 'Almacén':

            fechaS=request.vars.fecha_sumi.split("-")
            fecha_sumi=datetime.datetime(int(fechaS[0]),int(fechaS[1]),int(fechaS[2]))
            fechaComp=datetime.date(int(fechaS[0]),int(fechaS[1]),int(fechaS[2]))

            
            fechaHoy = datetime.datetime.now()
            if  fechaHoy < fecha_sumi:
                response.flash = "Fecha de ingreso no puede ser mayor a la actual"
                return False
            elif (fecha_sumi.year==fechaHoy.year and fecha_sumi.month == fechaHoy.month and\
                    fechaHoy.day-6 > 0 and fecha_sumi.day < (fechaHoy.day-6)):
                response.flash = "Fecha de ingreso no puede menor a una semana"
                return False    
            elif (fecha_sumi.year==fechaHoy.year and fecha_sumi.month != fechaHoy.month):
                response.flash = "Los registros se cierran mensualmente"
                return False           
    

            #############################################################
            #############################################################
            # Si hay registros con fechas siguientes a la que estoy ingresando 
            # se debe hacer un recalculo de los totales de los registros en esas fechas
            #  
            

            almacen = int(request.vars.almacen)

            db.t_Balance.insert(
                f_cantidad=cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto=concepto,
                f_tipo_ingreso=tipo_ing,
                f_fechaUso=fecha_sumi,
                f_medida=inv.f_medida,
                f_inventario=inv.id,
                f_sustancia=inv.sustancia,
                f_almacen=almacen)

        elif  tipo_ing == 'Prestamo':


            fechaS=request.vars.fecha_sumi.split("-")
            fecha_sumi=datetime.datetime(int(fechaS[0]),int(fechaS[1]),int(fechaS[2]))
            fechaComp=datetime.date(int(fechaS[0]),int(fechaS[1]),int(fechaS[2]))


            fechaHoy = datetime.datetime.now()
            if  fechaHoy < fecha_sumi:
                response.flash = "Fecha de ingreso no puede ser mayor a la actual"
                return False
            elif (fecha_sumi.year==fechaHoy.year and fecha_sumi.month == fechaHoy.month and\
                    fechaHoy.day-6 > 0 and fecha_sumi.day < (fechaHoy.day-6)):
                response.flash = "Fecha de ingreso no puede menor a una semana"
                return False
            elif (fecha_sumi.year==fechaHoy.year and fecha_sumi.month != fechaHoy.month):
                response.flash = "Los registros se cierran mensualmente"
                return False     
      
            db.t_Balance.insert(
                f_cantidad=cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto=concepto,
                f_tipo_ingreso=tipo_ing,
                f_fechaUso=fecha_sumi,
                f_medida=inv.f_medida,
                f_inventario=inv.id,
                f_sustancia=inv.sustancia)

        elif tipo_ing == 'Cesión':


            fechaS=request.vars.fecha_sumi.split("-")
            fecha_sumi=datetime.datetime(int(fechaS[0]),int(fechaS[1]),int(fechaS[2]))
            fechaComp=datetime.date(int(fechaS[0]),int(fechaS[1]),int(fechaS[2]))


            fechaHoy = datetime.datetime.now()
            if  fechaHoy < fecha_sumi:
                response.flash = "Fecha de ingreso no puede ser mayor a la actual"
                return False
            elif (fecha_sumi.year==fechaHoy.year and fecha_sumi.month == fechaHoy.month and\
                    fechaHoy.day-6 > 0 and fecha_sumi.day < (fechaHoy.day-6)):
                response.flash = "Fecha de ingreso no puede menor a una semana"
                return False
            elif (fecha_sumi.year==fechaHoy.year and fecha_sumi.month != fechaHoy.month):
                response.flash = "Los registros se cierran mensualmente"
                return False     

            db.t_Balance.insert(
                f_cantidad=cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto=concepto,
                f_tipo_ingreso=tipo_ing,
                f_fechaUso=fecha_sumi,
                f_medida=inv.f_medida,
                f_inventario=inv.id,
                f_sustancia=inv.sustancia)

        # Tipo ingreso es compra
        else:

            # Datos de la nueva compra
            nro_factura = request.vars.nro_factura
            institucion = request.vars.institucion
            rif = request.vars.rif

            fechaC=request.vars.fecha_compra.split("-")
            fecha_compra=datetime.datetime(int(fechaC[0]),int(fechaC[1]),int(fechaC[2]))
            fechaComp=datetime.date(int(fechaC[0]),int(fechaC[1]),int(fechaC[2]))


            fechaHoy = datetime.datetime.now()
            if fechaHoy < fecha_compra:
                response.flash = "Fecha de compra no puede ser mayor a la actual"
                return False 
            elif (fecha_compra.year==fechaHoy.year and fecha_compra.month == fechaHoy.month and\
                    fechaHoy.day-6 > 0 and fecha_compra.day < (fechaHoy.day-6)):
                response.flash = "Fecha de Compra no puede menor a una semana"
                return False
            elif (fecha_compra.year==fechaHoy.year and fecha_compra.month != fechaHoy.month):
                response.flash = "Los registros se cierran mensualmente"
                return False   

            # Se registra la nueva compra en la tabla t_Compra
            compra_id = db.t_Compra.insert(
                f_cantidad=cantidad,
                f_nro_factura=nro_factura,
                f_institucion=institucion,
                f_rif=rif,
                f_fecha=fecha_compra,
                f_sustancia=inv.sustancia,
                f_medida=unidad_id)

            db.t_Balance.insert(
                f_cantidad=cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto=concepto,
                f_tipo_ingreso=tipo_ing,
                f_medida=inv.f_medida,
                f_compra=compra_id,
                f_inventario=inv.id,
                f_sustancia=inv.sustancia,
                f_fechaUso=fecha_compra)

    # Si es un tipo Egreso 
    
    else:

        total_nuevo = total_viejo - cantidad
        uso_interno_nuevo = uso_interno_viejo - cantidad

        tipo_eg = request.vars.tipo_egreso  
        fecha_uso=request.vars.fecha_compra.split("-")
        fecha_u=datetime.date(int(fecha_uso[0]),int(fecha_uso[1]),int(fecha_uso[2]))  
        
                
        fechaHoy = datetime.datetime.now()
        if fechaHoy < fecha_u:
            response.flash = "Fecha de consumo no puede ser mayor a la actual"
            return False
        elif (fecha_u.year==fechaHoy.year and fecha_u.month == fechaHoy.month and\
                    fechaHoy.day-6 > 0 and fecha_u.day < (fechaHoy.day-6)):
                response.flash = "Fecha de Consumo no puede menor a una semana"
                return False 
        elif (fecha_u.year==fechaHoy.year and fecha_u.month != fechaHoy.month):
                response.flash = "Los registros se cierran mensualmente"
                return False  
        
        fechaComp=datetime.date(int(fecha_uso[0]),int(fecha_uso[1]),int(fecha_uso[2]))


        # Nueva cantidad total luego del consumo
        if total_nuevo <= 0:
            response.flash = "La cantidad total luego del consumo no puede ser "\
                             "negativa"
            return False        
        # Nueva cantidad de uso interno nueva puede ser maximo lo que era antes
        # (si hay material suficiente) o el nuevo total
        uso_interno_nuevo = min(uso_interno_viejo, total_nuevo)

        # Actualizando cantidad total con la nueva 
        inv.update_record(
            f_existencia=total_nuevo,
            f_uso_interno=uso_interno_nuevo)

        servicio_id = request.vars.servicio

        db.t_Balance.insert(
            f_cantidad=cantidad,
            f_cantidad_total=total_nuevo,
            f_fechaUso= fecha_uso,         
            f_concepto=concepto,
            f_tipo_egreso=tipo_eg,
            f_medida=inv.f_medida,
            f_servicio=servicio_id,
            f_inventario=inv.id,
            f_sustancia=inv.sustancia)

    # Se redirije para evitar mensaje de revisita con metodo POST
    return redirect(URL(args=request.args, vars=request.get_vars, host=True))

# Muestra los movimientos de la bitacora comenzando por el mas reciente
@auth.requires(lambda: __check_role())
@auth.requires_login(otherwise=URL('modulos', 'login'))
def bitacora():
    # INICIO Datos del modal de agregar un registro
    # Conceptos
    conceptos = ['Ingreso','Consumo']

    # Tipos de consumos
    #tipos_egreso = db.t_Balance.f_tipo_egreso.requires.other.theset
    tipos_egreso = ['Docencia','Investigación','Extensión']

    # Tipos de ingresos
    #tipos_ingreso = db.t_Balance.f_tipo_ingreso.requires.other.theset
    tipos_ingreso = ['Compra','Almacén']

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    # Lista de almacences
    almacenes = db(db.espacios_fisicos.id > 0).select()

    # Lista de servicios
    servicios = db(db.servicios.id > 0).select()
    # FIN Datos del modal de agregar un registro

    # Obteniendo la entrada en t_Personal del usuario conectado

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]

    if request.vars.inv is None:
        redirect(URL('inventarios'))

    inventario_id = int(request.vars.inv)

    # Si el id de inventario no es valido, retornar al inventario del
    # espacio fisico que se estaba consultando
    if not __is_valid_id(inventario_id, db.t_Inventario):
        response.flash = "La bitácora consultada no es correcta."
        redirect(URL('inventarios'))

    # Inventario al que pertenecen los registros que se desean consultar
    inventario = db((db.t_Inventario.id == inventario_id) & 
                    (db.t_Inventario.espacio == db.espacios_fisicos.id) & 
                    (db.t_Inventario.sustancia == db.t_Sustancia.id)
                   ).select()[0]

    # Espacio al que pertenece la bitacora consultada
    espacio_id = inventario['t_Inventario'].espacio

    # Unidad de medida en que es expresada la sustancia en el inventario
    unidad_medida = db(db.t_Unidad_de_medida.id == inventario.t_Inventario.f_medida
                      ).select()[0]

    # Se valida que el usuario tenga acceso a la bitacora indicada
    # para consultar la bitacora. 
    if not __acceso_permitido(user, espacio_id, "True"):
        redirect(URL('inventarios'))

    sust_nombre = inventario['t_Sustancia'].f_nombre

    espacio_nombre = inventario['espacios_fisicos'].codigo

    bitacora = db((db.t_Balance.f_inventario == inventario_id) &
                  (db.t_Balance.created_by == db.auth_user.id) &
                  (db.auth_user.id == db.t_Personal.f_usuario) &
                  (db.t_Balance.f_medida == db.t_Unidad_de_medida.id)).select()
    
    # *!* Hacer esto cuando se cree el registro y ponerlo en reg['f_descripcion']
    # Obteniendo la descripcion de cada fila y guardandola como un atributo
    for reg in bitacora:
        descripcion = __get_descripcion(reg['t_Balance'])
        reg['t_Balance']['descripcion'] = descripcion

    # Si se han enviado datos para agregar un nuevo registro
    concepto = request.vars.concepto
    if concepto:
        __agregar_registro(concepto)


    return dict(bitacora=bitacora,
                unidad_medida=unidad_medida,
                inventario=inventario,
                sust_nombre=sust_nombre,
                espacio_nombre=espacio_nombre,
                espacio_id=espacio_id,
                conceptos=conceptos,
                tipos_egreso=tipos_egreso,
                tipos_ingreso=tipos_ingreso,
                unidades_de_medida=unidades_de_medida,
                almacenes=almacenes,
                servicios=servicios)

# Muestra el inventario de acuerdo al cargo del usuario y la dependencia que tiene
# a cargo
@auth.requires(lambda: __check_role())
@auth.requires_login(otherwise=URL('modulos', 'login'))
def sustancia():

    # Inicializando listas de espacios fisicos y dependencias

    # OJO: Espacios debe ser [] siempre que no se este visitando un espacio fisico
    espacios = []
    dependencias = []
    dep_nombre = ""
    dep_padre_id = ""
    dep_padre_nombre = ""

    # Lista de sustancias en el inventario de un espacio fisico o que componen 
    # el inventario agregado de una dependencia
    inventario = []
    
    sust = request.vars.sust


    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    # Esta variable es enviada a la vista para que cuando el usuario seleccione 
    # un espacio fisico, se pase por GET es_espacio = "True". No quiere decir
    # que la dependencia seleccionada sea un espacio, sino que la siguiente
    # dependencia visitada sera un espacio fisico
    es_espacio = False

    # Permite saber si actualmente se esta visitando un espacio fisico (True)
    # o una dependencia (False)
    espacio_visitado = False

    # Indica si se debe seguir mostrando la flecha para seguir retrocediendo 
    retroceder = True
    
    es_tecnico = auth.has_membership("TÉCNICO")
    direccion_id = __find_dep_id('DIRECCIÓN')

    # Obteniendo la entrada en t_Personal del usuario conectado
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_id = user.id
    user_dep_id = user.f_dependencia

    if auth.has_membership("TÉCNICO"):
        # Si el tecnico ha seleccionado un espacio fisico
        if request.vars.dependencia:
            if request.vars.es_espacio == "True":
                # Evaluando la correctitud de los parametros del GET 
                if not (__is_valid_id(request.vars.dependencia, db.espacios_fisicos) and
                        __is_bool(request.vars.es_espacio)):
                    redirect(URL('inventarios'))

                # Determinando si el usuario tiene privilegios suficientes para
                # consultar la dependencia en request.vars.dependencia
                if not __acceso_permitido(user, 
                                    int(request.vars.dependencia), 
                                        request.vars.es_espacio):
                    redirect(URL('inventarios'))

                espacio_id = request.vars.dependencia
                espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                dep_nombre = espacio.codigo

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = espacio.dependencia
                dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                    ).select().first().nombre

                espacio_visitado = True

                # Busca el inventario del espacio
                inventario = __consultar_espacios(espacio_id, sust)

                sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

                # Si se esta agregando una nueva sustancia, se registra en la DB
                if request.vars.sustancia:
                    __agregar_sustancia(espacio,
                                        request.vars.sustancia, 
                                        request.vars.total,
                                        request.vars.uso_interno,
                                        request.vars.unidad)
            else:
                # Espacios a cargo del usuario user_id que pertenecen a la seccion
                # en request.vars.dependencia
                espacios = [row.espacios_fisicos for row in db(
                    (db.es_encargado.espacio_fisico == db.espacios_fisicos.id) & 
                    (db.espacios_fisicos.dependencia == int(request.vars.dependencia)) & 
                    (db.es_encargado.tecnico == user_id)).select()]

                espacios_ids = [e.id for e in espacios]

                dep_id = int(request.vars.dependencia)
                dep_nombre = db(db.dependencias.id == dep_id).select()[0].nombre

                dep_padre_nombre = "Secciones"

                # Se suman los inventarios de los espacios que tiene a cargo el usuario en la
                # seccion actual
                inventario = __consultar_espacios(espacios_ids, sust)

                es_espacio = True

        # Si el tecnico o jefe no ha seleccionado un espacio sino que acaba de 
        # entrar a la opcion de inventarios
        else:
            # Se buscan las secciones a las que pertenecen los espacios que
            # tiene a cargo el usuario
            espacios_a_cargo = db(
                (db.es_encargado.tecnico == user_id) & 
                (db.espacios_fisicos.id == db.es_encargado.espacio_fisico)
                                 ).select()

            secciones_ids = {e.espacios_fisicos.dependencia for e in espacios_a_cargo}

            dependencias = map(lambda x: db(db.dependencias.id == x).select()[0], 
                               secciones_ids)

            dep_nombre = "Secciones"

            espacios_ids = [e.espacios_fisicos.id for e in espacios_a_cargo]

            inventario = __consultar_espacios(espacios_ids, sust)

    elif auth.has_membership("JEFE DE SECCIÓN"):
        # Si el jefe de seccion ha seleccionado un espacio fisico
        if request.vars.es_espacio == 'True':
            # Determinando si el usuario tiene privilegios suficientes para
            # consultar la dependencia en request.vars.dependencia
            if not __acceso_permitido(user, 
                                int(request.vars.dependencia), 
                                    request.vars.es_espacio):
                redirect(URL('inventarios'))

            # Evaluando la correctitud de los parametros del GET 
            if not (__is_valid_id(request.vars.dependencia, db.espacios_fisicos) and
                    __is_bool(request.vars.es_espacio)):
                redirect(URL('inventarios'))


            espacio_id = request.vars.dependencia
            espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
            dep_nombre = db(db.espacios_fisicos.id == request.vars.dependencia
                           ).select().first().codigo

            # Guardando el ID y nombre de la dependencia a la que pertenece el 
            # espacio fisico visitado
            dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                             ).select().first().dependencia
            dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                 ).select().first().nombre

            espacio_visitado = True
                            # Se muestra la lista de sustancias que tiene en inventario
            inventario = __consultar_espacios(espacio_id, sust)

            sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

            # Si se esta agregando una nueva sustancia, se registra en la DB
            if request.vars.sustancia:
                __agregar_sustancia(espacio,
                                    request.vars.sustancia, 
                                    request.vars.total,
                                    request.vars.uso_interno,
                                    request.vars.unidad)


        # Si el jefe de seccion no ha seleccionado un espacio sino que acaba de 
        # regresar a la vista inicial de inventarios
        elif request.vars.es_espacio == 'False':
            if not (__is_valid_id(request.vars.dependencia, db.espacios_fisicos) and
                    __is_bool(request.vars.es_espacio)):
                    redirect(URL('inventarios'))
            # Determinando si el usuario tiene privilegios suficientes para
            # consultar la dependencia en request.vars.dependencia
            if not __acceso_permitido(user, 
                                int(request.vars.dependencia), 
                                    request.vars.es_espacio):
                redirect(URL('inventarios'))
            espacios = list(db(
                              db.espacios_fisicos.dependencia == user_dep_id
                              ).select(db.espacios_fisicos.ALL))
            dep_nombre = db(db.dependencias.id == user_dep_id
                           ).select().first().nombre

            es_espacio = True                        
        # Si el jefe de seccion no ha seleccionado un espacio sino que acaba de 
        # entrar a la vista inicial de inventarios
        else:
            espacios = list(db(
                              db.espacios_fisicos.dependencia == user_dep_id
                              ).select(db.espacios_fisicos.ALL))
            dep_nombre = db(db.dependencias.id == user_dep_id
                           ).select().first().nombre

            es_espacio = True

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la seccion del jefe
            inventario = __consultar_espacios(user_dep_id, sust)

    # Si el usuario no es tecnico, para la base de datos es indiferente su ROL
    # pues la jerarquia de dependencias esta almacenada en la misma tabla
    # con una lista de adyacencias
    else:
        # Si el usuario ha seleccionado una dependencia o un espacio fisico
        if request.vars.dependencia:

            # Evaluando la correctitud de los parametros del GET 
            if not (__is_valid_id(request.vars.dependencia, db.dependencias) and
                    __is_bool(request.vars.es_espacio)):
                redirect(URL('inventarios'))

            # Determinando si el usuario tiene privilegios suficientes para
            # consultar la dependencia en request.vars.dependencia
            if not __acceso_permitido(user, 
                                int(request.vars.dependencia), 
                                    request.vars.es_espacio):
                redirect(URL('inventarios'))

            if request.vars.es_espacio == "True":
        
                # Se muestra el inventario del espacio
                espacio_id = request.vars.dependencia
                espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                dep_nombre = espacio.codigo

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                                    ).select().first().dependencia
                dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                    ).select().first().nombre

                espacio_visitado = True

                # Se muestra la lista de sustancias que tiene en inventario
                inventario = __consultar_espacios(espacio_id, sust)

                sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

                # Si se esta agregando una nueva sustancia, se registra en la DB
                if request.vars.sustancia:
                    __agregar_sustancia(espacio,
                                        request.vars.sustancia, 
                                        request.vars.total,
                                        request.vars.uso_interno,
                                        request.vars.unidad)

            else:
                # Se muestran las dependencias que componen a esta dependencia padre
                # y se lista el inventario agregado
                dep_id = request.vars.dependencia
                dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                      ).select(db.dependencias.ALL))
                # Si la lista de dependencias es vacia, entonces la dependencia no 
                # tiene otras dependencias por debajo (podria tener espacios fisicos
                # o estar vacia)
                if len(dependencias) == 0:
                    # Buscando espacios fisicos que apunten a la dependencia escogida
                    espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                      ).select(db.espacios_fisicos.ALL))
                    es_espacio = True

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                 ).select().first().unidad_de_adscripcion
                # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                # y no hay un padre de este nodo
                if dep_padre_id:
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                         ).select().first().nombre
                # Se muestra como inventario el egregado de los inventarios que
                # pertenecen a la dependencia del usuario
                inventario = __solicitar_inventario_dep(dep_id, sust)

        else:
            # Dependencia a la que pertenece el usuario o que tiene a cargo
            dep_id = user.f_dependencia
            dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre

            # Se muestran las dependencias que componen a la dependencia que
            # tiene a cargo el usuario y el inventario agregado de esta
            dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                  ).select(db.dependencias.ALL))

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la dependencia del usuario
            inventario = __solicitar_inventario_dep(dep_id, sust)

    return dict(dep_nombre=dep_nombre, 
                dependencias=dependencias, 
                espacios=espacios, 
                es_espacio=es_espacio,
                espacio_visitado=espacio_visitado,
                dep_padre_id=dep_padre_id,
                dep_padre_nombre=dep_padre_nombre,
                direccion_id=direccion_id,
                es_tecnico=es_tecnico,
                inventario=inventario,
                unidades_de_medida=unidades_de_medida,
                retroceder=retroceder
                )


# Muestra el inventario de acuerdo al cargo del usuario y la dependencia que tiene
# a cargo
@auth.requires(lambda: __check_role())
@auth.requires_login(otherwise=URL('modulos', 'login'))
def inventarios():

    # Inicializando listas de espacios fisicos y dependencias

    # OJO: Espacios debe ser [] siempre que no se este visitando un espacio fisico
    espacios = []
    dependencias = []
    dep_nombre = ""
    dep_padre_id = ""
    dep_padre_nombre = ""

    # Lista de sustancias en el inventario de un espacio fisico o que componen 
    # el inventario agregado de una dependencia
    inventario = []
    
    # Lista de sustancias en el catalogo para el modal de agregar sustancia
    # al alcanzar el nivel de espacios fisicos
    sustancias = []

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    # Esta variable es enviada a la vista para que cuando el usuario seleccione 
    # un espacio fisico, se pase por GET es_espacio = "True". No quiere decir
    # que la dependencia seleccionada sea un espacio, sino que la siguiente
    # dependencia visitada sera un espacio fisico
    es_espacio = False

    # Permite saber si actualmente se esta visitando un espacio fisico (True)
    # o una dependencia (False)
    espacio_visitado = False
    
    # Indica si se debe seguir mostrando la flecha para seguir retrocediendo 
    retroceder = True

    es_tecnico = auth.has_membership("TÉCNICO")
    direccion_id = __find_dep_id('DIRECCIÓN')

    # Obteniendo la entrada en t_Personal del usuario conectado
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_id = user.id
    user_dep_id = user.f_dependencia

    if auth.has_membership("TÉCNICO"):
        # Si el tecnico ha seleccionado un espacio fisico
        if request.vars.dependencia:
            if request.vars.es_espacio == "True":
                # Evaluando la correctitud de los parametros del GET 
                if not (__is_valid_id(request.vars.dependencia, db.espacios_fisicos) and
                        __is_bool(request.vars.es_espacio)):
                    redirect(URL('inventarios'))

                # Determinando si el usuario tiene privilegios suficientes para
                # consultar la dependencia en request.vars.dependencia
                if not __acceso_permitido(user, 
                                    int(request.vars.dependencia), 
                                        request.vars.es_espacio):
                    redirect(URL('inventarios'))

                espacio_id = request.vars.dependencia
                espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                dep_nombre = espacio.codigo

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = espacio.dependencia
                dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                    ).select().first().nombre

                espacio_visitado = True

                # Busca el inventario del espacio
                inventario = __get_inventario_espacio(espacio_id)

                sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

                # Si se esta agregando una nueva sustancia, se registra en la DB
                if request.vars.sustancia:
                    __agregar_sustancia(espacio,
                                        request.vars.sustancia, 
                                        request.vars.total,
                                        request.vars.uso_interno,
                                        request.vars.unidad)
            else:
                # Espacios a cargo del usuario user_id que pertenecen a la seccion
                # en request.vars.dependencia
                espacios = [row.espacios_fisicos for row in db(
                    (db.es_encargado.espacio_fisico == db.espacios_fisicos.id) & 
                    (db.espacios_fisicos.dependencia == int(request.vars.dependencia)) & 
                    (db.es_encargado.tecnico == user_id)).select()]

                espacios_ids = [e.id for e in espacios]

                dep_id = int(request.vars.dependencia)
                dep_nombre = db(db.dependencias.id == dep_id).select()[0].nombre

                dep_padre_nombre = "Secciones"

                # Se suman los inventarios de los espacios que tiene a cargo el usuario en la
                # seccion actual
                inventario = __sumar_inventarios(espacios_ids)

                es_espacio = True

        # Si el tecnico o jefe no ha seleccionado un espacio sino que acaba de 
        # entrar a la opcion de inventarios
        else:
            # Se buscan las secciones a las que pertenecen los espacios que
            # tiene a cargo el usuario
            espacios_a_cargo = db(
                (db.es_encargado.tecnico == user_id) & 
                (db.espacios_fisicos.id == db.es_encargado.espacio_fisico)
                                 ).select()

            secciones_ids = {e.espacios_fisicos.dependencia for e in espacios_a_cargo}

            dependencias = map(lambda x: db(db.dependencias.id == x).select()[0], 
                               secciones_ids)

            dep_nombre = "Secciones"

            espacios_ids = [e.espacios_fisicos.id for e in espacios_a_cargo]

            inventario = __sumar_inventarios(espacios_ids)

    elif auth.has_membership("JEFE DE SECCIÓN"):
        # Si el jefe de seccion ha seleccionado un espacio fisico
        if request.vars.es_espacio == 'True':
            # Determinando si el usuario tiene privilegios suficientes para
            # consultar la dependencia en request.vars.dependencia
            if not __acceso_permitido(user, 
                                int(request.vars.dependencia), 
                                    request.vars.es_espacio):
                redirect(URL('inventarios'))

            # Evaluando la correctitud de los parametros del GET 
            if not (__is_valid_id(request.vars.dependencia, db.espacios_fisicos) and
                    __is_bool(request.vars.es_espacio)):
                redirect(URL('inventarios'))


            espacio_id = request.vars.dependencia
            espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
            dep_nombre = db(db.espacios_fisicos.id == request.vars.dependencia
                           ).select().first().codigo

            # Guardando el ID y nombre de la dependencia a la que pertenece el 
            # espacio fisico visitado
            dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                             ).select().first().dependencia
            dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                 ).select().first().nombre

            espacio_visitado = True
                            # Se muestra la lista de sustancias que tiene en inventario
            inventario = __get_inventario_espacio(espacio_id)

            sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

            # Si se esta agregando una nueva sustancia, se registra en la DB
            if request.vars.sustancia:
                __agregar_sustancia(espacio,
                                    request.vars.sustancia, 
                                    request.vars.total,
                                    request.vars.uso_interno,
                                    request.vars.unidad)


        # Si el jefe de seccion no ha seleccionado un espacio sino que acaba de 
        # regresar a la vista inicial de inventarios
        elif request.vars.es_espacio == 'False':
            if not (__is_valid_id(request.vars.dependencia, db.espacios_fisicos) and
                    __is_bool(request.vars.es_espacio)):
                    redirect(URL('inventarios'))
            # Determinando si el usuario tiene privilegios suficientes para
            # consultar la dependencia en request.vars.dependencia
            if not __acceso_permitido(user, 
                                int(request.vars.dependencia), 
                                    request.vars.es_espacio):
                redirect(URL('inventarios'))
            espacios = list(db(
                              db.espacios_fisicos.dependencia == user_dep_id
                              ).select(db.espacios_fisicos.ALL))
            dep_nombre = db(db.dependencias.id == user_dep_id
                           ).select().first().nombre

            es_espacio = True                        
        # Si el jefe de seccion no ha seleccionado un espacio sino que acaba de 
        # entrar a la vista inicial de inventarios
        else:
            espacios = list(db(
                              db.espacios_fisicos.dependencia == user_dep_id
                              ).select(db.espacios_fisicos.ALL))
            dep_nombre = db(db.dependencias.id == user_dep_id
                           ).select().first().nombre

            es_espacio = True

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la seccion del jefe
            inventario = __get_inventario_dep(user_dep_id)

    # Si el usuario no es tecnico, para la base de datos es indiferente su ROL
    # pues la jerarquia de dependencias esta almacenada en la misma tabla
    # con una lista de adyacencias
    else:
        # Si el usuario ha seleccionado una dependencia o un espacio fisico
        if request.vars.dependencia:

            # Evaluando la correctitud de los parametros del GET 
            if not (__is_valid_id(request.vars.dependencia, db.dependencias) and
                    __is_bool(request.vars.es_espacio)):
                redirect(URL('inventarios'))

            # Determinando si el usuario tiene privilegios suficientes para
            # consultar la dependencia en request.vars.dependencia
            if not __acceso_permitido(user, 
                                int(request.vars.dependencia), 
                                    request.vars.es_espacio):
                redirect(URL('inventarios'))

            if request.vars.es_espacio == "True":
        
                # Se muestra el inventario del espacio
                espacio_id = request.vars.dependencia
                espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                dep_nombre = espacio.codigo

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                                    ).select().first().dependencia
                dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                    ).select().first().nombre

                espacio_visitado = True

                # Se muestra la lista de sustancias que tiene en inventario
                inventario = __get_inventario_espacio(espacio_id)

                sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

                # Si se esta agregando una nueva sustancia, se registra en la DB
                if request.vars.sustancia:
                    __agregar_sustancia(espacio,
                                        request.vars.sustancia, 
                                        request.vars.total,
                                        request.vars.uso_interno,
                                        request.vars.unidad)

            else:
                # Se muestran las dependencias que componen a esta dependencia padre
                # y se lista el inventario agregado
                dep_id = request.vars.dependencia
                dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                      ).select(db.dependencias.ALL))
                # Si la lista de dependencias es vacia, entonces la dependencia no 
                # tiene otras dependencias por debajo (podria tener espacios fisicos
                # o estar vacia)
                if len(dependencias) == 0:
                    # Buscando espacios fisicos que apunten a la dependencia escogida
                    espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                      ).select(db.espacios_fisicos.ALL))
                    es_espacio = True

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                 ).select().first().unidad_de_adscripcion
                # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                # y no hay un padre de este nodo
                if dep_padre_id:
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                         ).select().first().nombre
                # Se muestra como inventario el egregado de los inventarios que
                # pertenecen a la dependencia del usuario
                inventario = __get_inventario_dep(dep_id)

        else:
            # Dependencia a la que pertenece el usuario o que tiene a cargo
            dep_id = user.f_dependencia
            dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre

            # Se muestran las dependencias que componen a la dependencia que
            # tiene a cargo el usuario y el inventario agregado de esta
            dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                  ).select(db.dependencias.ALL))

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la dependencia del usuario
            inventario = __get_inventario_dep(dep_id)

    return dict(dep_nombre=dep_nombre, 
                dependencias=dependencias, 
                espacios=espacios, 
                es_espacio=es_espacio,
                espacio_visitado=espacio_visitado,
                dep_padre_id=dep_padre_id,
                dep_padre_nombre=dep_padre_nombre,
                direccion_id=direccion_id,
                es_tecnico=es_tecnico,
                inventario=inventario,
                sustancias=sustancias,
                unidades_de_medida=unidades_de_medida,
                retroceder=retroceder)

########################################
#         ENVASES/CONTENEDORES         #
# FUNCIONES AUXILIARES Y CONTROLADORES #
########################################
@auth.requires_login(otherwise=URL('modulos', 'login'))
def envases():
    user_id = auth.user_id

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    categorias_de_desecho = []
    contenedores = []
    espacios_fisicos_adscritos = []
    unidades_de_medida = []

    formas = ['Cilíndrica', 'Cuadrada', 'Rectangular', 'Otra']
    materiales = ['Plástico', 'Polietileno (HDPE)', 'Polietileno (PE)', 'Vidrio', 'Metal', 'Acero', 'Otro']
    tipos_de_boca = ['Boca ancha', 'Boca angosta', 'Cerrados con abertura de trasvase', 'Otra']

    categorias_de_desecho = list(db(db.t_categoria_desechos).select(db.t_categoria_desechos.ALL))

    # Listas de espacios físicos de los cuáles el usuario logueado es responsable
    # Si el usuario es un gestor o webmaster, puede crear envases en cualquier espacio físico
    if(auth.has_membership('GESTOR DE SMyDP') or  auth.has_membership('WEBMASTER') or auth.has_membership('DIRECTOR') or auth.has_membership('ASISTENTE DEL DIRECTOR') or auth.has_membership("PERSONAL INTERNO")):
        contenedores = list(db(db.t_envases).select(db.t_envases.ALL))
        espacios_fisicos_adscritos = list(db(
            (db.espacios_fisicos)
        ).select(db.espacios_fisicos.id, db.espacios_fisicos.codigo)) 

    elif(auth.has_membership('TÉCNICO') or auth.has_membership('JEFE DE LABORATORIO')):
        #pero si no es un gestor o webmaster, solamente puede crear contenedores en los espacios físicos en donde tiene
        #jurisdicción
        contenedores = list(db(
            (db.t_envases.espacio_fisico == db.espacios_fisicos.id) &
            (db.espacios_fisicos.dependencia == db.dependencias.id ) &
            (db.dependencias.unidad_de_adscripcion == user_dep_id) 
        ).select(db.t_envases.ALL))

        espacios_fisicos_adscritos = list(db(
            (db.espacios_fisicos.dependencia == db.dependencias.id) &
            (db.dependencias.unidad_de_adscripcion == user_dep_id) 
        ).select(db.espacios_fisicos.id, db.espacios_fisicos.codigo)) 

    elif auth.has_membership('JEFE DE SECCIÓN'):
        #pero si no es un gestor o webmaster, solamente puede crear contenedores en los espacios físicos en donde tiene
        #jurisdicción
        contenedores = list(db(
            (db.t_envases.espacio_fisico == db.espacios_fisicos.id) &
            (db.espacios_fisicos.dependencia == user_dep_id)
        ).select(db.t_envases.ALL))

        espacios_fisicos_adscritos = list(db(
            (db.espacios_fisicos.dependencia == user_dep_id)
        ).select(db.espacios_fisicos.id, db.espacios_fisicos.codigo)) 


    unidades_de_medida = list(db(db.t_Unidad_de_medida).select(db.t_Unidad_de_medida.ALL))

    # El formulario de edición/creación de un envase se ha recibido
    if request.vars.capacidad:

        envase = {}

        marcado_para_borrar = False

        if request.vars.borrar_envase == 'True':
            marcado_para_borrar = True

        # Verifica si el elemento fue marcado para ser borrado
        if marcado_para_borrar:
            response.flash = __eliminar_envase(int(request.vars.id_envase))
            session.flash = response.flash
            return redirect(URL(host=True)) 

        else:
            #De lo contrario debe ser creado o actualizado
            id_envase = -1

            if request.vars.id_envase != '':
                id_envase = int(request.vars.id_envase)
            
            response.flash = __agregar_envase(
                request.vars.identificacion,
                float(str(request.vars.capacidad).replace(",", ".")),
                int(request.vars.unidad_medida),
                request.vars.forma,
                request.vars.material,
                request.vars.tipo_boca,
                request.vars.descripcion,
                request.vars.composicion,
                int(request.vars.espacio_fisico),
                int(request.vars.categoria),
                id_envase
            )

            session.flash = response.flash
            return redirect(URL(host=True)) 

    return locals()


def __agregar_envase(identificacion, capacidad, unidad_medida, forma, material, tipo_boca, descripcion, composicion, espacio_fisico, categoria, id_envase):
    # Si el id_envase es distinto de -1, es porque ya existe el envase y se va a actualizar su informacion
    if id_envase != -1:
        if len(list(db((db.t_envases.identificacion == identificacion) & (db.t_envases.id != id_envase)).select())) > 0:
            return T("La identificación que proporcionó para el contenedor ya se encuentra en uso.")
        else:
            db(db.t_envases.id == id_envase).update(
                identificacion = identificacion.upper(),
                capacidad = capacidad, 
                unidad_medida = unidad_medida, 
                forma = forma.upper(), 
                material = material.upper(), 
                tipo_boca = tipo_boca.upper(), 
                descripcion = descripcion.upper(), 
                composicion = composicion, 
                espacio_fisico = espacio_fisico, 
                categoria = categoria
            )

            return T("La información del contenedor se ha modificado exitosamente.")
        
    else:
        # Se verifica si la identificación del envase que se quiere crear fue previamente utilizada
        if len(list(db(db.t_envases.identificacion == identificacion).select())) > 0:
            return T("La identificación que proporcionó para el contenedor ya se encuentra en uso.")
            
        else:
            #De lo contrario, el envase aún no existe y se tiene que crear
            db.t_envases.insert(
                identificacion = identificacion.upper(),
                capacidad = capacidad, 
                unidad_medida = unidad_medida, 
                forma = forma.upper(), 
                material = material.upper(), 
                tipo_boca = tipo_boca.upper(), 
                descripcion = descripcion.upper(), 
                composicion = composicion, 
                espacio_fisico = espacio_fisico, 
                categoria = categoria
            )

            return T("Contenedor creado exitosamente.")

def __eliminar_desecho(id_desecho):
    db(db.t_inventario_desechos.id == id_desecho).delete()
    return T("Desecho eliminado exitosamente.")

def __eliminar_envase(id_envase):
    db(db.t_envases.id == id_envase).delete()
    return T("Contenedor eliminado exitosamente.")

########################################
#         CATEGORIAS DE DESECHOS       #
# FUNCIONES AUXILIARES Y CONTROLADORES #
########################################
@auth.requires_login(otherwise=URL('modulos', 'login'))
def categorias_desechos():
    categorias = []

    if(auth.has_membership('GESTOR DE SMyDP') or  auth.has_membership('WEBMASTER')):
        categorias = list(db(db.t_categoria_desechos
                                  ).select(db.t_categoria_desechos.ALL))
        # El formulario de edición/creación de categoria se ha recibido
        if request.vars.categoria:
            marcado_para_borrar = False

            if request.vars.borrar_categoria == 'True':
                marcado_para_borrar = True

            # Verifica si el elemento fue marcado para ser borrado
            if marcado_para_borrar:
                
                response.flash = __eliminar_categoria(int(request.vars.id_categoria))
                session.flash = response.flash
                return redirect(URL(host=True)) 
            else:
                #De lo contrario debe ser creado o actualizado
                id_categoria = -1

                if request.vars.id_categoria != '':
                    id_categoria = int(request.vars.id_categoria)
                
                response.flash = __agregar_categoria(request.vars.categoria, request.vars.descripcion, id_categoria)
                session.flash = response.flash
                return redirect(URL(host=True)) 

    else:
        categorias = list(db(db.t_categoria_desechos
                                  ).select(db.t_categoria_desechos.ALL))
    return locals()


def __agregar_categoria(nombre_categoria, descripcion_categoria, id_categoria):
    # Si el id_categoria es distinto de -1, es porque ya exista la categoría y se va a actualizar
    if id_categoria != -1:
        db(db.t_categoria_desechos.id == id_categoria).update(categoria = nombre_categoria.upper(), descripcion = descripcion_categoria.upper())
    else:
        # Se verifica que la categoría no exista previamente
        if(len(list(db(db.t_categoria_desechos.categoria == nombre_categoria).select())) > 0):
            return T("La categoría que intenta agregar ya existe")
        else:
            #De lo contrario, la categoría no existe y se tiene que crear
            db.t_categoria_desechos.insert(categoria = nombre_categoria.upper(), descripcion = descripcion_categoria.upper())
            return T("Categoría agregada exitosamente")
            



def __eliminar_categoria(categoria_id):
    db(db.t_categoria_desechos.id == categoria_id).delete()
    return T("Categoría de desecho eliminada exitosamente.")


# @auth.requires(lambda: __check_role())
@auth.requires_login(otherwise=URL('modulos', 'login'))
def inventarios_desechos():
    # Inicializando listas de espacios fisicos y dependencias

    # OJO: Espacios debe ser [] siempre que no se este visitando un espacio fisico
    espacios = []
    dependencias = []
    dep_nombre = ""
    dep_padre_id = ""
    dep_padre_nombre = ""
    envases = []
    envases_totales = []
    inventario_total = []
    # Este valor indica si se muestra el campo "Dependencia" en la tabla del inventario 
    # si se esta visitando el menu principal del inventario, o si se está visitando una sección
    mostrar_campo_dependencia = False 

    # Lista de sustancias en el inventario de un espacio fisico o que componen 
    # el inventario agregado de una dependencia
    inventario = []
    
    # Lista de sustancias en el catalogo para el modal de agregar sustancia
    # al alcanzar el nivel de espacios fisicos
    desechos = []

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    # Esta variable es enviada a la vista para que cuando el usuario seleccione 
    # un espacio fisico, se pase por GET es_espacio = "True". No quiere decir
    # que la dependencia seleccionada sea un espacio, sino que la siguiente
    # dependencia visitada sera un espacio fisico
    es_espacio = False

    # Permite saber si actualmente se esta visitando un espacio fisico (True)
    # o una dependencia (False)
    espacio_visitado = False
    
    # Indica si se debe seguir mostrando la flecha para seguir retrocediendo 
    retroceder = True

    es_tecnico = auth.has_membership("TÉCNICO")
    direccion_id = __find_dep_id('DIRECCIÓN')

    # Obteniendo la entrada en t_Personal del usuario conectado
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_id = user.id
    user_dep_id = user.f_dependencia

    if(auth.has_membership('GESTOR DE SMyDP') or  auth.has_membership('WEBMASTER') or auth.has_membership('DIRECTOR') or auth.has_membership('ASISTENTE DEL DIRECTOR')):
        # Si el usuario ha seleccionado una dependencia o un espacio fisico
        if request.vars.dependencia:
            if request.vars.es_espacio == "True":
                # Se muestra el inventario del espacio
                espacio_id = request.vars.dependencia
                espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                dep_nombre = espacio.codigo

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                                    ).select().first().dependencia
                dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                    ).select().first().nombre

                espacio_visitado = True

                # Se muestra la lista de sustancias que tiene en inventario

                inventario = list(db(
                    (db.espacios_fisicos.id == espacio_id) &
                    (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                    (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.composicion, 
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.responsable,
                    db.t_inventario_desechos.envase,
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.peligrosidad,
                    db.t_inventario_desechos.tratamiento,
                    groupby = 
                     db.t_inventario_desechos.categoria |  
                    db.t_inventario_desechos.id | 
                     db.t_inventario_desechos.composicion | 
                     db.t_inventario_desechos.responsable |
                     db.t_inventario_desechos.unidad_medida |
                     db.t_inventario_desechos.peligrosidad |
                     db.t_inventario_desechos.tratamiento |
                     db.t_inventario_desechos.envase
                ))

                inventario_total = list(db(
                    (db.espacios_fisicos.id == espacio_id) &
                    (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                    (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                    ).select())

                ####################
                # A T E N C I Ó N  #
                ####################
                # Cuando se va a subir el sistema a produccion, descomentar la linea que dice "t_bitacora_desecho" y comentar la que dice "t_Bitacora_desecho"
                # Analogamente, comentar la línea correcta cuando se está en ambiente de desarrollo
                envases = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ' and e.id not in (select entrada.envase from "t_Bitacora_desechos" entrada);', as_dict = True))
                #envases = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ' and e.id not in (select entrada.envase from "t_bitacora_desechos" entrada);', as_dict = True))

                envases_totales = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ';', as_dict = True))
                
                # Se quiere eliminar un desecho
               # print request.vars
                if request.vars.view and request.vars.borrar_desecho:
                    marcado_para_borrar = False
                    ##print request.vars
                    if request.vars.borrar_desecho == 'True':
                       # print "marcado para borrar = true"
                        marcado_para_borrar = True

                    # Verifica si el elemento fue marcado para ser borrado
                    if marcado_para_borrar:
                       # print "se va a borrar"
                        response.flash = __eliminar_desecho(int(request.vars.view))
                       # print "ya se tuvo que haber borrado"
                        session.flash = response.flash
                        return redirect(URL('..', 'sigulab2', 'smydp/inventarios_desechos', vars=dict(dependencia=request.vars.dependencia, es_espacio="True"))) 


                # Se esta editando el detalle de un desecho
                if request.vars.view and request.vars.envase:
                    envase = db(db.t_envases.id == int(request.vars.envase)).select().first()
                    
                    response.flash = __actualizar_desecho(request.vars.view, envase, request.vars.peligrosidad, request.vars.tratamiento, request.vars.concentracion)
                    session.flash = response.flash
                    return redirect(URL('..', 'sigulab2', 'smydp/inventarios_desechos', vars=dict(dependencia=request.var.dependencia, es_espacio="True", view=request.vars.view))) 
                
                else:
                    # Si se esta agregando un nuevo desecho, se registra en la DB
                    if request.vars.envase:

                        #Se busca la información del envase en la DB
                        envase = list(db(db.t_envases.id == request.vars.envase).select())

                        response.flash =__agregar_desecho(envase[0],
                                            request.vars.peligrosidad,
                                            request.vars.tratamiento,
                                            request.vars.cantidad,
                                            request.vars.concentracion
                        )

                        session.flash = response.flash
                        return redirect(URL(host=True)) 

            else:
                # Se muestran las dependencias que componen a esta dependencia padre
                # y se lista el inventario agregado
                dep_id = request.vars.dependencia
                dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                      ).select(db.dependencias.ALL))
                # Si la lista de dependencias es vacia, entonces la dependencia no 
                # tiene otras dependencias por debajo (podria tener espacios fisicos
                # o estar vacia)
                
                if len(dependencias) == 0:
                    # Buscando espacios fisicos que apunten a la dependencia escogida
                    espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                      ).select(db.espacios_fisicos.ALL))
                    
                    inventario = list(db(
                        (db.t_inventario_desechos.seccion == request.vars.dependencia)).select())

                    inventario = list(db(
                    (db.dependencias.id == dep_id) & 
                    (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                    (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.espacio_fisico,
                    db.t_inventario_desechos.seccion,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.responsable,
                    groupby = 
                     db.t_inventario_desechos.categoria |
                     db.t_inventario_desechos.espacio_fisico |
                        db.t_inventario_desechos.id | 
                    db.t_inventario_desechos.seccion |
                    db.t_inventario_desechos.unidad_medida | 
                    db.t_inventario_desechos.responsable
                    ))

                    es_espacio = True
                
                else:
                    inventario = list(db(
                    (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.dependencias.unidad_de_adscripcion == request.vars.dependencia)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.espacio_fisico,
                    db.t_inventario_desechos.seccion,
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.responsable,
                    groupby = 
                     db.t_inventario_desechos.categoria |
                        db.t_inventario_desechos.id | 
                     db.t_inventario_desechos.espacio_fisico |
                     db.t_inventario_desechos.seccion |
                    db.t_inventario_desechos.unidad_medida | 
                    db.t_inventario_desechos.responsable
                    ))

                    mostrar_campo_dependencia = True

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                 ).select().first().unidad_de_adscripcion
                # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                # y no hay un padre de este nodo
                if dep_padre_id:
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                         ).select().first().nombre


        else:
            # Dependencia a la que pertenece el usuario o que tiene a cargo
            dep_id = user.f_dependencia
            dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre

            # Se muestran las dependencias que componen a la dependencia que
            # tiene a cargo el usuario y el inventario agregado de esta
            dependencias = list(db(db.dependencias.nombre.startswith('LAB')).select(db.dependencias.ALL))

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la dependencia del usuario
            inventario = list(db(
                (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                (db.espacios_fisicos.dependencia == db.dependencias.id)
                ).select(
                db.t_inventario_desechos.categoria,
                db.t_inventario_desechos.espacio_fisico,
                db.t_inventario_desechos.id,
                db.t_inventario_desechos.seccion,
                db.t_inventario_desechos.cantidad.sum(),
                db.t_inventario_desechos.unidad_medida,
                db.t_inventario_desechos.responsable,
                groupby = 
                    db.t_inventario_desechos.categoria |
                    db.t_inventario_desechos.espacio_fisico |
                    db.t_inventario_desechos.id | 
                    db.t_inventario_desechos.seccion |
                db.t_inventario_desechos.unidad_medida | 
                db.t_inventario_desechos.responsable
                ))

            mostrar_campo_dependencia = True

    elif(auth.has_membership("TÉCNICO") or auth.has_membership("JEFE DE LABORATORIO")):
        # Si el usuario ha seleccionado una dependencia o un espacio fisico
        if request.vars.dependencia:
            if request.vars.es_espacio == "True":
                # Si se está editando un desecho, se actualiza la informacion
                if request.vars.es_espacio == "True":
                    # Se muestra el inventario del espacio
                    espacio_id = request.vars.dependencia
                    espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                    dep_nombre = espacio.codigo

                    # Guardando el ID y nombre de la dependencia padre para el link 
                    # de navegacion de retorno
                    dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                                        ).select().first().dependencia
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                        ).select().first().nombre

                    espacio_visitado = True

                    # Se muestra la lista de sustancias que tiene en inventario

                    inventario = list(db(
                        (db.espacios_fisicos.id == espacio_id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                        ).select(
                        db.t_inventario_desechos.categoria,
                        db.t_inventario_desechos.id,
                        db.t_inventario_desechos.composicion, 
                        db.t_inventario_desechos.cantidad.sum(),
                        db.t_inventario_desechos.responsable,
                        db.t_inventario_desechos.envase,
                        db.t_inventario_desechos.unidad_medida,
                        db.t_inventario_desechos.peligrosidad,
                        db.t_inventario_desechos.tratamiento,
                        groupby = 
                        db.t_inventario_desechos.categoria |  
                        db.t_inventario_desechos.id | 
                        db.t_inventario_desechos.composicion | 
                        db.t_inventario_desechos.responsable |
                        db.t_inventario_desechos.unidad_medida |
                        db.t_inventario_desechos.peligrosidad |
                        db.t_inventario_desechos.tratamiento |
                        db.t_inventario_desechos.envase
                    ))

                    inventario_total = list(db(
                        (db.espacios_fisicos.id == espacio_id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                        ).select())

                    ####################
                    # A T E N C I Ó N  #
                    ####################
                    # Cuando se va a subir el sistema a produccion, descomentar la linea que dice "t_bitacora_desecho" y comentar la que dice "t_Bitacora_desecho"
                    # Analogamente, comentar la línea correcta cuando se está en ambiente de desarrollo
                    envases = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ' and e.id not in (select entrada.envase from "t_Bitacora_desechos" entrada);', as_dict = True))
                    #envases = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ' and e.id not in (select entrada.envase from "t_bitacora_desechos" entrada);', as_dict = True))

                    envases_totales = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ';', as_dict = True))
                    
                    # Se esta editando el detalle de un desecho
                    if request.vars.view and request.vars.envase:
                        envase = db(db.t_envases.id == int(request.vars.envase)).select().first()
                        
                        response.flash = __actualizar_desecho(request.vars.view, envase, request.vars.peligrosidad, request.vars.tratamiento, request.vars.concentracion)
                        session.flash = response.flash
                        return redirect(URL(host=True)) 
                    else:
                        # Si se esta agregando un nuevo desecho, se registra en la DB
                        if request.vars.envase:

                            #Se busca la información del envase en la DB
                            envase = list(db(db.t_envases.id == request.vars.envase).select())

                            response.flash = __agregar_desecho(envase[0],
                                                request.vars.peligrosidad,
                                                request.vars.tratamiento,
                                                request.vars.cantidad,
                                                request.vars.concentracion
                            )

                            session.flash = response.flash
                            return redirect(URL(host=True)) 

                else:
                    # Se muestran las dependencias que componen a esta dependencia padre
                    # y se lista el inventario agregado
                    dep_id = request.vars.dependencia
                    dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                    dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                        ).select(db.dependencias.ALL))                                                                                                                                                                              
                    # Si la lista de dependencias es vacia, entonces la dependencia no 
                    # tiene otras dependencias por debajo (podria tener espacios fisicos                                                        
                    # o estar vacia)

                    if len(dependencias) == 0:
                        # Buscando espacios fisicos que apunten a la dependencia escogida
                        espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                        ).select(db.espacios_fisicos.ALL))
                        
                        inventario = list(db(
                            (db.t_inventario_desechos.seccion == request.vars.dependencia)).select())

                        inventario = list(db(
                        (db.dependencias.id == dep_id) & 
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                        ).select(
                        db.t_inventario_desechos.categoria,
                        db.t_inventario_desechos.espacio_fisico,
                        db.t_inventario_desechos.seccion,
                        db.t_inventario_desechos.id,
                        db.t_inventario_desechos.cantidad.sum(),
                        db.t_inventario_desechos.unidad_medida,
                        db.t_inventario_desechos.responsable,
                        groupby = 
                        db.t_inventario_desechos.categoria |
                        db.t_inventario_desechos.espacio_fisico |
                            db.t_inventario_desechos.id | 
                        db.t_inventario_desechos.seccion |
                        db.t_inventario_desechos.unidad_medida | 
                        db.t_inventario_desechos.responsable
                        ))

                        es_espacio = True
                    
                    else:
                        inventario = list(db(
                        (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                            (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                            (db.dependencias.unidad_de_adscripcion == request.vars.dependencia)
                        ).select(
                        db.t_inventario_desechos.categoria,
                        db.t_inventario_desechos.id,
                        db.t_inventario_desechos.espacio_fisico,
                        db.t_inventario_desechos.seccion,
                        db.t_inventario_desechos.cantidad.sum(),
                        db.t_inventario_desechos.unidad_medida,
                        db.t_inventario_desechos.responsable,
                        groupby = 
                        db.t_inventario_desechos.categoria |
                            db.t_inventario_desechos.id | 
                        db.t_inventario_desechos.espacio_fisico |
                        db.t_inventario_desechos.seccion |
                        db.t_inventario_desechos.unidad_medida | 
                        db.t_inventario_desechos.responsable
                        ))

                        mostrar_campo_dependencia = True

                    # Guardando el ID y nombre de la dependencia padre para el link 
                    # de navegacion de retorno
                    dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                    ).select().first().unidad_de_adscripcion
                    # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                    # y no hay un padre de este nodo
                    if dep_padre_id:
                        dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                            ).select().first().nombre

            else:
                # Se muestran las dependencias que componen a esta dependencia padre
                # y se lista el inventario agregado
                dep_id = request.vars.dependencia
                dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                      ).select(db.dependencias.ALL))
                # Si la lista de dependencias es vacia, entonces la dependencia no 
                # tiene otras dependencias por debajo (podria tener espacios fisicos
                # o estar vacia)
                
                if len(dependencias) == 0:
                    # Buscando espacios fisicos que apunten a la dependencia escogida
                    espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                      ).select(db.espacios_fisicos.ALL))
                    
                    inventario = list(db(
                        (db.t_inventario_desechos.seccion == request.vars.dependencia)).select())

                    inventario = list(db(
                    (db.dependencias.id == dep_id) & 
                    (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                    (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.espacio_fisico,
                    db.t_inventario_desechos.seccion,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.responsable,
                    groupby = 
                     db.t_inventario_desechos.categoria |
                     db.t_inventario_desechos.espacio_fisico |
                        db.t_inventario_desechos.id | 
                    db.t_inventario_desechos.seccion |
                    db.t_inventario_desechos.unidad_medida | 
                    db.t_inventario_desechos.responsable
                    ))


                    es_espacio = True
                
                else:
                    inventario = list(db(
                    (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.dependencias.unidad_de_adscripcion == request.vars.dependencia)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.espacio_fisico,
                    db.t_inventario_desechos.seccion,
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.responsable,
                    groupby = 
                     db.t_inventario_desechos.categoria |
                        db.t_inventario_desechos.id | 
                     db.t_inventario_desechos.espacio_fisico |
                     db.t_inventario_desechos.seccion |
                    db.t_inventario_desechos.unidad_medida | 
                    db.t_inventario_desechos.responsable
                    ))

                    mostrar_campo_dependencia = True

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                 ).select().first().unidad_de_adscripcion
                # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                # y no hay un padre de este nodo
                if dep_padre_id:
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                         ).select().first().nombre


        else:
            # Dependencia a la que pertenece el usuario o que tiene a cargo
            dep_id = user.f_dependencia
            dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre

            # Se muestran las dependencias que componen a la dependencia que
            # tiene a cargo el usuario y el inventario agregado de esta
            dependencias = list(db(
                (db.dependencias.unidad_de_adscripcion == dep_id)
            ).select(db.dependencias.ALL))

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la dependencia del usuario

            inventario = list(db(
                (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                (db.espacios_fisicos.dependencia == db.dependencias.id) &
                (db.dependencias.unidad_de_adscripcion == dep_id)
                ).select(
                db.t_inventario_desechos.categoria,
                db.t_inventario_desechos.espacio_fisico,
                db.t_inventario_desechos.id,
                db.t_inventario_desechos.seccion,
                db.t_inventario_desechos.cantidad.sum(),
                db.t_inventario_desechos.unidad_medida,
                db.t_inventario_desechos.responsable,
                groupby = 
                    db.t_inventario_desechos.categoria |
                    db.t_inventario_desechos.espacio_fisico |
                    db.t_inventario_desechos.id | 
                    db.t_inventario_desechos.seccion |
                db.t_inventario_desechos.unidad_medida | 
                db.t_inventario_desechos.responsable
                ))

            mostrar_campo_dependencia = True
            espacio_visitado = False

    elif auth.has_membership("JEFE DE SECCIÓN") or auth.has_membership("PERSONAL INTERNO") :
        # Si el usuario ha seleccionado una dependencia o un espacio fisico
        if request.vars.dependencia:
            if request.vars.es_espacio == "True":
                # Si se está editando un desecho, se actualiza la informacion
                if request.vars.es_espacio == "True":
                    # Se muestra el inventario del espacio
                    espacio_id = request.vars.dependencia
                    espacio = db(db.espacios_fisicos.id == espacio_id).select()[0]
                    dep_nombre = espacio.codigo

                    # Guardando el ID y nombre de la dependencia padre para el link 
                    # de navegacion de retorno
                    dep_padre_id = db(db.espacios_fisicos.id == request.vars.dependencia
                                        ).select().first().dependencia
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                        ).select().first().nombre

                    espacio_visitado = True

                    # Se muestra la lista de sustancias que tiene en inventario

                    inventario = list(db(
                        (db.espacios_fisicos.id == espacio_id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                        ).select(
                        db.t_inventario_desechos.categoria,
                        db.t_inventario_desechos.id,
                        db.t_inventario_desechos.composicion, 
                        db.t_inventario_desechos.cantidad.sum(),
                        db.t_inventario_desechos.responsable,
                        db.t_inventario_desechos.envase,
                        db.t_inventario_desechos.unidad_medida,
                        db.t_inventario_desechos.peligrosidad,
                        db.t_inventario_desechos.tratamiento,
                        groupby = 
                        db.t_inventario_desechos.categoria |  
                        db.t_inventario_desechos.id | 
                        db.t_inventario_desechos.composicion | 
                        db.t_inventario_desechos.responsable |
                        db.t_inventario_desechos.unidad_medida |
                        db.t_inventario_desechos.peligrosidad |
                        db.t_inventario_desechos.tratamiento |
                        db.t_inventario_desechos.envase
                    ))

                    inventario_total = list(db(
                        (db.espacios_fisicos.id == espacio_id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                        ).select())

                    ####################
                    # A T E N C I Ó N  #
                    ####################
                    # Cuando se va a subir el sistema a produccion, descomentar la linea que dice "t_bitacora_desecho" y comentar la que dice "t_Bitacora_desecho"
                    # Analogamente, comentar la línea correcta cuando se está en ambiente de desarrollo
                    envases = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ' and e.id not in (select entrada.envase from "t_Bitacora_desechos" entrada);', as_dict = True))
                    #envases = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ' and e.id not in (select entrada.envase from "t_bitacora_desechos" entrada);', as_dict = True))

                    envases_totales = list(db.executesql('SELECT * from t_envases e where e.espacio_fisico = ' + espacio_id + ';', as_dict = True))
                    
                    # Se esta editando el detalle de un desecho
                    if request.vars.view and request.vars.envase:
                        envase = db(db.t_envases.id == int(request.vars.envase)).select().first()
                        
                        response.flash = __actualizar_desecho(request.vars.view, envase, request.vars.peligrosidad, request.vars.tratamiento, request.vars.concentracion)
                        session.flash = response.flash
                        return redirect(URL(host=True)) 
                    else:
                        # Si se esta agregando un nuevo desecho, se registra en la DB
                        if request.vars.envase:

                            #Se busca la información del envase en la DB
                            envase = list(db(db.t_envases.id == request.vars.envase).select())

                            response.flash = __agregar_desecho(envase[0],
                                                request.vars.peligrosidad,
                                                request.vars.tratamiento,
                                                request.vars.cantidad,
                                                request.vars.concentracion
                            )

                            session.flash = response.flash
                            return redirect(URL(host=True)) 

                else:
                    # Se muestran las dependencias que componen a esta dependencia padre
                    # y se lista el inventario agregado
                    dep_id = request.vars.dependencia
                    dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                    dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                        ).select(db.dependencias.ALL))
                    # Si la lista de dependencias es vacia, entonces la dependencia no 
                    # tiene otras dependencias por debajo (podria tener espacios fisicos
                    # o estar vacia)
                    
                    if len(dependencias) == 0:
                        # Buscando espacios fisicos que apunten a la dependencia escogida
                        espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                        ).select(db.espacios_fisicos.ALL))
                        
                        inventario = list(db(
                            (db.t_inventario_desechos.seccion == request.vars.dependencia)).select())

                        inventario = list(db(
                        (db.dependencias.id == dep_id) & 
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                        ).select(
                        db.t_inventario_desechos.categoria,
                        db.t_inventario_desechos.espacio_fisico,
                        db.t_inventario_desechos.seccion,
                        db.t_inventario_desechos.id,
                        db.t_inventario_desechos.cantidad.sum(),
                        db.t_inventario_desechos.unidad_medida,
                        db.t_inventario_desechos.responsable,
                        groupby = 
                        db.t_inventario_desechos.categoria |
                        db.t_inventario_desechos.espacio_fisico |
                            db.t_inventario_desechos.id | 
                        db.t_inventario_desechos.seccion |
                        db.t_inventario_desechos.unidad_medida | 
                        db.t_inventario_desechos.responsable
                        ))

                        es_espacio = True
                    
                    else:
                        inventario = list(db(
                        (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                            (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                            (db.dependencias.unidad_de_adscripcion == request.vars.dependencia)
                        ).select(
                        db.t_inventario_desechos.categoria,
                        db.t_inventario_desechos.id,
                        db.t_inventario_desechos.espacio_fisico,
                        db.t_inventario_desechos.seccion,
                        db.t_inventario_desechos.cantidad.sum(),
                        db.t_inventario_desechos.unidad_medida,
                        db.t_inventario_desechos.responsable,
                        groupby = 
                        db.t_inventario_desechos.categoria |
                            db.t_inventario_desechos.id | 
                        db.t_inventario_desechos.espacio_fisico |
                        db.t_inventario_desechos.seccion |
                        db.t_inventario_desechos.unidad_medida | 
                        db.t_inventario_desechos.responsable
                        ))

                        mostrar_campo_dependencia = True

                    # Guardando el ID y nombre de la dependencia padre para el link 
                    # de navegacion de retorno
                    dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                    ).select().first().unidad_de_adscripcion
                    # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                    # y no hay un padre de este nodo
                    if dep_padre_id:
                        dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                            ).select().first().nombre

            else:
                # Se muestran las dependencias que componen a esta dependencia padre
                # y se lista el inventario agregado
                dep_id = request.vars.dependencia
                dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre
                dependencias = list(db(db.dependencias.unidad_de_adscripcion == dep_id
                                      ).select(db.dependencias.ALL))
                # Si la lista de dependencias es vacia, entonces la dependencia no 
                # tiene otras dependencias por debajo (podria tener espacios fisicos
                # o estar vacia)
                
                if len(dependencias) == 0:
                    # Buscando espacios fisicos que apunten a la dependencia escogida
                    espacios = list(db(db.espacios_fisicos.dependencia == dep_id
                                      ).select(db.espacios_fisicos.ALL))
                    
                    inventario = list(db(
                        (db.t_inventario_desechos.seccion == request.vars.dependencia)).select())

                    inventario = list(db(
                    (db.dependencias.id == dep_id) & 
                    (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                    (db.espacios_fisicos.id == db.t_inventario_desechos.espacio_fisico)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.espacio_fisico,
                    db.t_inventario_desechos.seccion,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.responsable,
                    groupby = 
                     db.t_inventario_desechos.categoria |
                     db.t_inventario_desechos.espacio_fisico |
                        db.t_inventario_desechos.id | 
                    db.t_inventario_desechos.seccion |
                    db.t_inventario_desechos.unidad_medida | 
                    db.t_inventario_desechos.responsable
                    ))


                    es_espacio = True
                
                else:
                    inventario = list(db(
                    (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                        (db.espacios_fisicos.dependencia == db.dependencias.id) & 
                        (db.dependencias.unidad_de_adscripcion == request.vars.dependencia)
                    ).select(
                    db.t_inventario_desechos.categoria,
                    db.t_inventario_desechos.id,
                    db.t_inventario_desechos.espacio_fisico,
                    db.t_inventario_desechos.seccion,
                    db.t_inventario_desechos.cantidad.sum(),
                    db.t_inventario_desechos.unidad_medida,
                    db.t_inventario_desechos.responsable,
                    groupby = 
                     db.t_inventario_desechos.categoria |
                        db.t_inventario_desechos.id | 
                     db.t_inventario_desechos.espacio_fisico |
                     db.t_inventario_desechos.seccion |
                    db.t_inventario_desechos.unidad_medida | 
                    db.t_inventario_desechos.responsable
                    ))

                    mostrar_campo_dependencia = True

                # Guardando el ID y nombre de la dependencia padre para el link 
                # de navegacion de retorno
                dep_padre_id = db(db.dependencias.id == request.vars.dependencia
                                 ).select().first().unidad_de_adscripcion
                # Si dep_padre_id es None, se ha llegado al tope de la jerarquia
                # y no hay un padre de este nodo
                if dep_padre_id:
                    dep_padre_nombre = db(db.dependencias.id == dep_padre_id
                                         ).select().first().nombre


        else:
            # Dependencia a la que pertenece el usuario o que tiene a cargo
            dep_id = user.f_dependencia
            dep_nombre = db.dependencias(db.dependencias.id == dep_id).nombre

            # Se muestran las dependencias que componen a la dependencia que
            # tiene a cargo el usuario y el inventario agregado de esta
            dependencias = list(db(
                (db.dependencias.id == dep_id)
            ).select(db.dependencias.ALL))

            # Se muestra como inventario el egregado de los inventarios que
            # pertenecen a la dependencia del usuario

            inventario = list(db(
                (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id) &
                (db.espacios_fisicos.dependencia == db.dependencias.id) &
                (db.dependencias.id == dep_id)
                ).select(
                db.t_inventario_desechos.categoria,
                db.t_inventario_desechos.espacio_fisico,
                db.t_inventario_desechos.id,
                db.t_inventario_desechos.seccion,
                db.t_inventario_desechos.cantidad.sum(),
                db.t_inventario_desechos.unidad_medida,
                db.t_inventario_desechos.responsable,
                groupby = 
                    db.t_inventario_desechos.categoria |
                    db.t_inventario_desechos.espacio_fisico |
                    db.t_inventario_desechos.id | 
                    db.t_inventario_desechos.seccion |
                db.t_inventario_desechos.unidad_medida | 
                db.t_inventario_desechos.responsable
                ))

            mostrar_campo_dependencia = True
            espacio_visitado = False


    return dict(dep_nombre=dep_nombre, 
                dependencias=dependencias, 
                espacios=espacios, 
                es_espacio=es_espacio,
                espacio_visitado=espacio_visitado,
                dep_padre_id=dep_padre_id,
                dep_padre_nombre=dep_padre_nombre,
                direccion_id=direccion_id,
                es_tecnico=es_tecnico,
                inventario=inventario,
                desechos=desechos,
                envases=envases,
                envases_totales=envases_totales,
                inventario_total=inventario_total,
                unidades_de_medida=unidades_de_medida,
                retroceder=retroceder,
                mostrar_campo_dependencia = mostrar_campo_dependencia)


########################################
#          BITÁCORA DESECHOS           #
# FUNCIONES AUXILIARES Y CONTROLADORES #
########################################
def __agregar_entrada_bitacora_desechos(fecha, descripcion, cantidad_generada, cantidad_retirada, saldo_anterior, envase, inventario):
    fecha_parseada = datetime.datetime.strptime(fecha, "%Y-%m-%d")

    saldo_actual = float(cantidad_generada) + float(saldo_anterior) - float(cantidad_retirada)
    
    # Se actualiza la cantidad del desecho peligroso en la entrada del inventario
    row = db(db.t_inventario_desechos.id == int(inventario)).select().first()

    row.update_record(
        cantidad = saldo_actual
    )
    
    # Se agrega la entrada en la bitácora
    id = db.t_Bitacora_desechos.insert(
                fecha = fecha_parseada,
                descripcion = descripcion.upper(),
                cantidad_generada = cantidad_generada,
                cantidad_retirada = cantidad_retirada,
                saldo = saldo_actual,
                unidad_medida_bitacora = envase.unidad_medida,
                envase = envase.id,
                inventario = inventario
    )

    return T("Entrada agregada exitósamente.")

# Actualiza la información de un desecho peligroso
def __actualizar_desecho(id, envase, peligrosidad, tratamiento, concentracion):
    # Verifica que no existe en el inventario una entrada repetida 
    # se considera que una entrada es una única cuando una determinada composición
    # con una determinada unidad de medida ya se encuentra en un determinado especifico
    busqueda = 0

    busqueda = len(list(db(
        (db.t_inventario_desechos.composicion == envase.composicion) &
        (db.t_inventario_desechos.espacio_fisico == envase.espacio_fisico) &
        (db.t_inventario_desechos.unidad_medida == envase.unidad_medida) &
        (db.t_inventario_desechos.envase != envase.id) 
    ).select()))

    if busqueda == 0:
        row = db(db.t_inventario_desechos.id == int(id)).select().first()
        
        peligrosidad_mayusculas = []
        
        if type(peligrosidad) is list:
            peligrosidad_mayusculas = [x.upper() for x in peligrosidad]
        else:
            peligrosidad_mayusculas = peligrosidad.upper()
            
        row.update_record(
            categoria = envase.categoria,
            composicion = envase.composicion,
            envase = envase.id,
            concentracion = concentracion,
            espacio_fisico = envase.espacio_fisico,
            seccion = envase.espacio_fisico.dependencia,
            tratamiento = tratamiento.upper(),
            peligrosidad = peligrosidad_mayusculas
        )

        return T("Desecho actualizado correctamente.")

    else:
        return T("El desecho que usted está intentando ingresar ya se encuentra registrado. Por favor edite su entrada en la bitácora.")


# Agrega un nuevo desecho peligroso al inventario de un espacio físico
def __agregar_desecho(envase, peligrosidad, tratamiento, cantidad, concentracion):
    
    # Verifica que no existe en el inventario una entrada repetida 
    # se considera que una entrada es una única cuando una determinada composición
    # con una determinada unidad de medida ya se encuentra en un determinado especifico
    busqueda = 0
    busqueda = len(list(db(
        (db.t_inventario_desechos.composicion == envase.composicion) &
        (db.t_inventario_desechos.espacio_fisico == envase.espacio_fisico) &
        (db.t_inventario_desechos.unidad_medida == envase.unidad_medida) 
    ).select()))

    if busqueda == 0:
        # Verifica que la cantidad de desecho que se quiere registrar quepa dentro de la capacidad
        # del envase seleccionado
        if int(cantidad) <= int(envase.capacidad): 
            peligrosidad_mayusculas = []

            if type(peligrosidad) is list:
                peligrosidad_mayusculas = [x.upper() for x in peligrosidad]
            else:
                peligrosidad_mayusculas = peligrosidad.upper()

            #Agrega el desecho al inventario
            nueva_entrada_id = db.t_inventario_desechos.insert(categoria = envase.categoria,
                                            cantidad = cantidad,
                                            unidad_medida = envase.unidad_medida,
                                            composicion = envase.composicion,
                                            concentracion = concentracion,
                                            espacio_fisico = envase.espacio_fisico,
                                            seccion = envase.espacio_fisico.dependencia,
                                            responsable = auth.user_id,
                                            envase = envase.id,
                                            tratamiento = tratamiento.upper(),
                                            peligrosidad = peligrosidad_mayusculas)
            
            # Crea la entrada inicial en la bitácora de desechos
            db.t_Bitacora_desechos.insert(
                fecha = str(datetime.datetime.now()),
                descripcion = "ENTRADA INICIAL",
                cantidad_generada = cantidad,
                cantidad_retirada = 0,
                saldo = cantidad,
                unidad_medida_bitacora = envase.unidad_medida,
                envase = envase.id,
                inventario = nueva_entrada_id
            )

            return T("Desecho creado exitósamente.")


        else:
            return T("El contenedor que usted eligió no tiene la capacidad suficiente para almacenar la cantidad de desecho indicada.")
    else:
        return T("El desecho que usted está intentando ingresar ya se encuentra registrado. Por favor edite su entrada en la bitácora.")



# Muestra los movimientos de la bitacora comenzando por el mas reciente
@auth.requires(lambda: __check_role())
@auth.requires_login(otherwise=URL('modulos', 'login'))
def bitacora_desechos():
    # INICIO Datos del modal de agregar un registro
    # Conceptos
    conceptos = ['Generación','Retiro']

    # Tipos de consumos

    # Tipos de ingresos

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    # Obteniendo la entrada en t_Personal del usuario conectado

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]

    if request.vars.inv is None:
        redirect(URL('inventarios_desechos'))

    inventario_id = int(request.vars.inv)

    # Inventario al que pertenecen los registros que se desean consultar
    inventario = db((db.t_inventario_desechos.id == inventario_id) & 
                    (db.t_inventario_desechos.espacio_fisico == db.espacios_fisicos.id)
                   ).select()[0]
    
    # Espacio al que pertenece la bitacora consultada
    espacio_id = inventario['t_inventario_desechos']['espacio_fisico']['id']

    # Unidad de medida en que es expresada la sustancia en el inventario
    unidad_medida = inventario['t_inventario_desechos']['unidad_medida']['f_abreviatura']

    espacio_nombre = inventario['t_inventario_desechos']['espacio_fisico']['codigo']

    envases = list(db((db.t_envases.espacio_fisico == espacio_id)).select())

    bitacora = list(db(
        (db.t_Bitacora_desechos.inventario == inventario_id) & 
        (db.t_Bitacora_desechos.created_by == db.auth_user.id)
    ).select(orderby=~db.t_Bitacora_desechos.id))

    ultima_entrada = db(
        (db.t_Bitacora_desechos.inventario == inventario_id) &
        (db.t_Bitacora_desechos.created_by == db.auth_user.id)
    ).select(orderby=~db.t_Bitacora_desechos.id).first()

    #Se está agregando una nueva entrada a la bitácora
    if request.vars.fecha_entrada:
        cantidad_retirada = 0
        cantidad_generada = 0

        envase = db((db.t_envases.id == request.vars.envase)).select().first()

        if(request.vars.cantidad_retirada != ''):
            cantidad_retirada = request.vars.cantidad_retirada

        if(request.vars.cantidad_generada != ''):
            cantidad_generada = request.vars.cantidad_generada
        
        response.flash = __agregar_entrada_bitacora_desechos(
            request.vars.fecha_entrada,
            request.vars.descripcion,
            cantidad_generada,
            cantidad_retirada,
            ultima_entrada['t_Bitacora_desechos']['saldo'],
            envase,
            bitacora[0]['t_Bitacora_desechos']['inventario'].id
        )
        session.flash = response.flash
        
        return redirect(URL('..', 'sigulab2', 'smydp/bitacora_desechos', vars=dict(inv=inventario_id))) 


    return dict(bitacora=bitacora,
                unidad_medida=unidad_medida,
                inventario=inventario,
                composicion=inventario['t_inventario_desechos']['composicion'],
                espacio_nombre=espacio_nombre,
                espacio_id=espacio_id,
                conceptos=conceptos,
                unidades_de_medida=unidades_de_medida,
                envases = envases,
                fecha_actual = str(datetime.datetime.now()).split(" ")[0]
                )

@auth.requires_login(otherwise=URL('modulos', 'login'))
def desechos():
    return locals()



#--------------------- Catalogo de Sustancias y Materiales ----------

@auth.requires_login(otherwise=URL('modulos', 'login'))
def catalogo():

    if(auth.has_membership('GESTOR DE SMyDP') or  auth.has_membership('WEBMASTER')):
        table = SQLFORM.smartgrid(  db.t_Sustancia, 
                                    onupdate=auth.archive,
                                    links_in_grid=False,
                                    csv=False,
                                    user_signature=True,
                                    paginate=10)

    else:
        table = SQLFORM.smartgrid(  db.t_Sustancia, 
                                    editable=False,
                                    deletable=False,
                                    csv=False,
                                    links_in_grid=False,
                                    create=False,
                                    paginate=10)
    return locals()

@auth.requires_login(otherwise=URL('modulos', 'login'))
def detalles_solicitud():

    solicitud = db((db.t_Solicitud_smydp.f_cod_registro == request.vars.registro)).select()[0]

    sustancia = db((db.t_Sustancia.id == solicitud.f_sustancia)).select()[0]

    espacio = db((db.espacios_fisicos.id == solicitud.f_espacio)).select()[0]

    responsable = db(db.t_Personal.f_usuario == solicitud.f_responsable_solicitud).select()[0]

    medida = db(db.t_Unidad_de_medida.id == solicitud.f_medida).select()[0]


    #Posibles respuesta a la solicitud
    respuestas = ['Negación','Aceptación']

    #Posibles forma en que se tratara la solicitud
    formas = ['Cesión','Préstamo']

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    personal_usuario = db(auth.user_id == db.t_Personal.f_usuario).select(db.t_Personal.ALL)[0]

    # Espacios a cargo del usuario actual
    espacios = []
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    espacios_a_cargo = __get_espacios(user_dep_id)

    for esp in espacios_a_cargo:

        for row in db((db.t_Inventario.sustancia == solicitud.f_sustancia) &
                        (solicitud.f_espacio != esp) &
                        (db.t_Inventario.espacio == esp)
                        ).select():

            esp_aux = db((db.espacios_fisicos.id == esp)
                ).select()[0]

            espacios.append(esp_aux)

    dependencia_usuario = db(personal_usuario.f_dependencia == db.dependencias.id).select(db.dependencias.ALL)[0]

    registro = solicitud.f_cod_registro

    num_resp = validador_registro_respuestas(request, db, registro)

    nombre_dependencia = dependencia_usuario.nombre

    id_jefe_dependencia = dependencia_usuario.id_jefe_dependencia

    usuario_jefe = db(id_jefe_dependencia == auth.user.id).select(db.auth_user.ALL)[0]

    nombre_jefe = usuario_jefe.first_name
    apellido_jefe = usuario_jefe.last_name
    email_jefe = usuario_jefe.email

    nombre_responsable = personal_usuario.f_nombre
    email_responsable = personal_usuario.f_email
    esta_autorizado = not(auth.has_membership("TÉCNICO"))

    datos_solicitud = [nombre_dependencia, nombre_jefe, apellido_jefe, email_jefe, nombre_responsable, email_responsable, num_resp]



        #----- AGREGAR RESPUESTA -----#
    if request.post_vars.numResp:

        espacio = request.vars.espacio
        unidad = request.vars.unidad
        cantidad = float(request.vars.suministrar)
        invResp=db(db.t_Inventario.espacio==espacio).select()
        medidaInv = db(db.t_Unidad_de_medida.id == invResp[0].f_medida).select()[0]
        medidaResp = db(db.t_Unidad_de_medida.id == unidad).select()[0]
        cantidadAux = __transformar_cantidad(cantidad, medidaResp.f_nombre, medidaInv.f_nombre)

        if cantidadAux <= invResp[0].f_uso_interno:

            cantidad = float(request.vars.suministrar)
            unidad = request.vars.unidad
            respuesta = request.vars.respuesta
            sustancia = sustancia.id
            justificacion = request.vars.justificacion
            forma = request.vars.forma
            fecha_tope = request.vars.fecha_tope
            espacio = request.vars.espacio
            numResp = request.post_vars.numResp
            inv_id = db.t_Respuesta.insert(f_cod_registro=numResp, 
                                            f_cantidad= cantidad,
                                            f_medida=unidad, 
                                            f_tipo_respuesta=respuesta,
                                            f_justificacion=justificacion,
                                            f_calidad=forma,
                                            f_fecha_tope_devolucion=fecha_tope,
                                            f_espacio=espacio,
                                            f_solicitud=solicitud.id,
                                            f_responsable_entrega=personal_usuario.id)

        else:
            response.flash = "Cantidad sobrepasa lo que se tiene de uso interno"
            session.flash = response.flash
            redirect(URL(args=request.args, vars=request.get_vars, host=True)) 



        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

    return dict(solicitud = solicitud,
                sustancia = sustancia,
                medida = medida,
                espacio = espacio,
                datos_solicitud=datos_solicitud,
                responsable = responsable,
                respuestas = respuestas,
                formas = formas,
                esta_autorizado = esta_autorizado,
                espacios = espacios,
                unidades_de_medida=unidades_de_medida
                )


@auth.requires_login(otherwise=URL('modulos', 'login'))
def detalles_solicitud_realizada():

    solicitud = db((db.t_Solicitud_smydp.f_cod_registro == request.vars.registro)).select()[0]

    sustancia = db((db.t_Sustancia.id == solicitud.f_sustancia)).select()[0]

    espacio = db((db.espacios_fisicos.id == solicitud.f_espacio)).select()[0]

    responsable = db(db.t_Personal.f_usuario == solicitud.f_responsable_solicitud).select()[0]

    medida = db(db.t_Unidad_de_medida.id == solicitud.f_medida).select()[0]


    #Posibles respuesta a la solicitud
    respuestas = ['Negación','Aceptación']

    #Posibles forma en que se tratara la solicitud
    formas = ['Cesión','Préstamo']

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    personal_usuario = db(auth.user_id == db.t_Personal.f_usuario).select(db.t_Personal.ALL)[0]

    # Espacios a cargo del usuario actual
    espacios = []
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    espacios_a_cargo = __get_espacios(user_dep_id)

    for esp in espacios_a_cargo:

        for row in db((db.t_Inventario.sustancia == solicitud.f_sustancia) &
                        (db.t_Inventario.espacio == esp)).select():

            esp_aux = db((db.espacios_fisicos.id == esp)
                ).select()[0]

            espacios.append(esp_aux)

    dependencia_usuario = db(personal_usuario.f_dependencia == db.dependencias.id).select(db.dependencias.ALL)[0]

    registro = solicitud.f_cod_registro

    num_resp = validador_registro_respuestas(request, db, registro)

    nombre_dependencia = dependencia_usuario.nombre

    id_jefe_dependencia = dependencia_usuario.id_jefe_dependencia

    usuario_jefe = db(id_jefe_dependencia == auth.user.id).select(db.auth_user.ALL)[0]

    nombre_jefe = usuario_jefe.first_name
    apellido_jefe = usuario_jefe.last_name
    email_jefe = usuario_jefe.email

    nombre_responsable = personal_usuario.f_nombre
    email_responsable = personal_usuario.f_email
    esta_autorizado = not(auth.has_membership("TÉCNICO"))

    datos_solicitud = [nombre_dependencia, nombre_jefe, apellido_jefe, email_jefe, nombre_responsable, email_responsable, num_resp]

    return dict(solicitud = solicitud,
                sustancia = sustancia,
                medida = medida,
                espacio = espacio,
                datos_solicitud=datos_solicitud,
                responsable = responsable,
                respuestas = respuestas,
                formas = formas,
                esta_autorizado = esta_autorizado,
                espacios = espacios,
                unidades_de_medida=unidades_de_medida
                )


@auth.requires_login(otherwise=URL('modulos', 'login'))
def detalles_respuesta():

    respuesta = db((db.t_Respuesta.f_cod_registro == request.vars.registro)).select()[0]

    solicitud = db((db.t_Solicitud_smydp.id == respuesta.f_solicitud)).select()[0]

    sustancia = db((db.t_Sustancia.id == solicitud.f_sustancia)).select()[0]

    espacio = db((db.espacios_fisicos.id == respuesta.f_espacio)).select()[0]

    responsable = db(db.t_Personal.f_usuario == respuesta.f_responsable_entrega).select()[0]

    medida = db(db.t_Unidad_de_medida.id == respuesta.f_medida).select()[0]

    esta_autorizado = not(auth.has_membership("TÉCNICO"))

    #Posibles respuesta a la solicitud
    respuestas = ['Negación','Aceptación']

    #Posibles forma en que se tratara la solicitud
    formas = ['Cesión','Préstamo']

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    personal_usuario = db(auth.user_id == db.t_Personal.f_usuario).select(db.t_Personal.ALL)[0]

    # Espacios a cargo del usuario actual
    espacios = []
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    espacios_a_cargo = __get_espacios(user_dep_id)

    for esp in espacios_a_cargo:

        for row in db((db.t_Inventario.sustancia == solicitud.f_sustancia) &
                        (db.t_Inventario.espacio == esp)).select():

            esp_aux = db((db.espacios_fisicos.id == esp)
                ).select()[0]

            espacios.append(esp_aux)

    dependencia_usuario = db(personal_usuario.f_dependencia == db.dependencias.id).select(db.dependencias.ALL)[0]

    registro = solicitud.f_cod_registro

    num_resp = validador_registro_respuestas(request, db, registro)

    nombre_dependencia = dependencia_usuario.nombre

    id_jefe_dependencia = dependencia_usuario.id_jefe_dependencia

    usuario_jefe = db(id_jefe_dependencia == auth.user.id).select(db.auth_user.ALL)[0]

    nombre_jefe = usuario_jefe.first_name
    apellido_jefe = usuario_jefe.last_name
    email_jefe = usuario_jefe.email

    nombre_responsable = personal_usuario.f_nombre
    email_responsable = personal_usuario.f_email

    datos_solicitud = [nombre_dependencia, nombre_jefe, apellido_jefe, email_jefe, nombre_responsable, email_responsable, num_resp]
    
    
    #----- AGREGAR RESPUESTA -----#
    if request.post_vars.ci_receptor:
        sol = db((db.t_Solicitud_smydp.id == respuesta.f_solicitud)).select()[0]

        medidaResp = db(db.t_Unidad_de_medida.id == respuesta.f_medida).select()[0]
        medidaSol = db(db.t_Unidad_de_medida.id == sol.f_medida).select()[0]

        cantidad = __transformar_cantidad(respuesta.f_cantidad, medidaResp.f_nombre, medidaSol.f_nombre)

        if sol.f_cantidad_conseguida+cantidad <= sol.f_cantidad:

            ##Cargo la CI del receptor de la sustancia 

            cedula_receptor = request.post_vars.ci_receptor
            invRespo=db(db.t_Personal.f_ci==cedula_receptor).select()

            if len(invRespo)== 0:
                response.flash = "No existe un usuario con esta cedula"
                session.flash = response.flash
                return redirect(URL('respuestas'))

            inv_id = db(db.t_Respuesta.id == respuesta.id).update(f_responsable_recepcion = invRespo[0].id)


            #COMO ESTA SOLICITANDO TIENE UN INGRESO
            #Inventario de quien hace la solicitud calculado por sustancia
            invSolic=db((db.t_Inventario.sustancia==sol.f_sustancia) & (db.t_Inventario.espacio==sol.f_espacio) ).select()

            if len(invSolic) == 0:
                inv_id = db.t_Inventario.insert(f_existencia=float(0), 
                                        f_uso_interno=float(0),
                                        f_medida=respuesta.f_medida,
                                        espacio=solicitud.f_espacio,
                                        sustancia=solicitud.f_sustancia)

                concepto = 'Ingreso'
                tipo_ing = 'Ingreso inicial'

                # Agregando la primera entrada de la sustancia en la bitacora
                db.t_Balance.insert(
                                        f_cantidad=float(0),
                                        f_cantidad_total=float(0),
                                        f_concepto=concepto,
                                        f_tipo_ingreso=tipo_ing,
                                        f_medida=respuesta.f_medida,
                                        f_inventario=inv_id,
                                        f_sustancia=solicitud.f_sustancia,
                                        f_fechaUso=datetime.date.today())

                invSolic=db((db.t_Inventario.sustancia==sol.f_sustancia) & (db.t_Inventario.espacio==sol.f_espacio) ).select()


            # Busco el balance de la sustancia a la que me voy a hacer referencia
            balSol=db(db.t_Balance.f_inventario== invSolic[0].id).select()
            #Busco la sustancia a la que estoy sacandole cuentas 
            
        
            total_viejo= invSolic[0].f_existencia
            uso_int_viejo=invSolic[0].f_uso_interno
            espacio = db(db.espacios_fisicos.id == respuesta.f_espacio).select()[0]

            medidaInv = db(db.t_Unidad_de_medida.id == invSolic[0].f_medida).select()[0]
            medidaResp = db(db.t_Unidad_de_medida.id == respuesta.f_medida).select()[0]

            cantidad = __transformar_cantidad(respuesta.f_cantidad, medidaResp.f_nombre, medidaInv.f_nombre)

            total_nuevo = total_viejo + cantidad
            uso_interno_nuevo = uso_int_viejo + cantidad
            
            invSolic[0].update_record(
                    f_existencia=total_nuevo,
                    f_uso_interno=uso_interno_nuevo)

            db.t_Balance.insert(
                f_cantidad = cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto='Ingreso',
                f_tipo_ingreso=respuesta.f_calidad[0],
                f_fechaUso=datetime.datetime.now,
                f_medida=invSolic[0].f_medida,
                f_inventario=invSolic[0].id,
                f_sustancia=solicitud.f_sustancia,
                f_respuesta_solicitud=respuesta.id
            )
        
            #COMO ESTA DANDO TIENE UN CONSUMO
            #Inventario de quien hace la solicitud calculado por sustancia
            invRespo=db((db.t_Inventario.sustancia==sol.f_sustancia) & (db.t_Inventario.espacio==respuesta.f_espacio) ).select()
            # Busco el balance de la sustancia a la que me voy a hacer referencia
            balResp=db(db.t_Balance.f_inventario== invRespo[0].id).select()
            #Busco la sustancia a la que estoy sacandole cuentas 
            
        
            total_viejo= invRespo[0].f_existencia
            uso_int_viejo=invRespo[0].f_uso_interno
            espacio = db(db.espacios_fisicos.id == respuesta.f_espacio).select()[0]

            medidaInv = db(db.t_Unidad_de_medida.id == invRespo[0].f_medida).select()[0]

            cantidad = __transformar_cantidad(respuesta.f_cantidad, medidaResp.f_nombre, medidaInv.f_nombre)

            total_nuevo = total_viejo - cantidad
            uso_interno_nuevo = uso_int_viejo - cantidad
            
            invRespo[0].update_record(
                    f_existencia=total_nuevo,
                    f_uso_interno=uso_interno_nuevo)

            db.t_Balance.insert(

                f_cantidad = float(cantidad),
                f_cantidad_total=total_nuevo,
                f_concepto='Consumo',
                f_tipo_egreso=respuesta.f_calidad[0],
                f_fechaUso=datetime.datetime.now,
                f_medida=invRespo[0].f_medida,
                f_inventario=invRespo[0].id,
                f_sustancia=sol.f_sustancia
            )



            medidaResp = db(db.t_Unidad_de_medida.id == respuesta.f_medida).select()[0]
            medidaSol = db(db.t_Unidad_de_medida.id == sol.f_medida).select()[0]

            cantidad = __transformar_cantidad(respuesta.f_cantidad, medidaResp.f_nombre, medidaSol.f_nombre)

            sol.update_record(f_cantidad_conseguida=sol.f_cantidad_conseguida+cantidad)

            if str(respuesta.f_calidad[0])== 'Préstamo':
                sol.update_record(
                        f_estatus='Prestamo por devolver')

            if sol.f_cantidad_conseguida == sol.f_cantidad:

                if str(sol.f_estatus[0]) != 'Prestamo por devolver':

                    sol.update_record(f_estatus='Completada')

        else :

            response.flash = "Cantidad sobrepasa lo solicitado"
            session.flash = response.flash


        return redirect(URL('respuestas'))

    return dict(respuesta = respuesta,
                medida = medida,
                solicitud = solicitud,
                sustancia = sustancia,
                espacio = espacio,
                datos_solicitud=datos_solicitud,
                responsable = responsable,
                respuestas = respuestas,
                formas = formas,
                esta_autorizado = esta_autorizado,
                espacios = espacios,
                unidades_de_medida=unidades_de_medida
                )

@auth.requires_login(otherwise=URL('modulos', 'login'))
def detalles_respuesta_realizada():

    respuesta = db((db.t_Respuesta.f_cod_registro == request.vars.registro)).select()[0]

    solicitud = db((db.t_Solicitud_smydp.id == respuesta.f_solicitud)).select()[0]

    sustancia = db((db.t_Sustancia.id == solicitud.f_sustancia)).select()[0]

    espacio = db((db.espacios_fisicos.id == respuesta.f_espacio)).select()[0]

    responsable = db(db.t_Personal.f_usuario == respuesta.f_responsable_entrega).select()[0]

    medida = db(db.t_Unidad_de_medida.id == respuesta.f_medida).select()[0]

    esta_autorizado = not(auth.has_membership("TÉCNICO"))

    #Posibles respuesta a la solicitud
    respuestas = ['Negación','Aceptación']

    #Posibles forma en que se tratara la solicitud
    formas = ['Cesión','Préstamo']

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    personal_usuario = db(auth.user_id == db.t_Personal.f_usuario).select(db.t_Personal.ALL)[0]

    # Espacios a cargo del usuario actual
    espacios = []
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    espacios_a_cargo = __get_espacios(user_dep_id)

    for esp in espacios_a_cargo:

        for row in db((db.t_Inventario.sustancia == solicitud.f_sustancia) &
                        (db.t_Inventario.espacio == esp)).select():

            esp_aux = db((db.espacios_fisicos.id == esp)
                ).select()[0]

            espacios.append(esp_aux)

    dependencia_usuario = db(personal_usuario.f_dependencia == db.dependencias.id).select(db.dependencias.ALL)[0]

    registro = solicitud.f_cod_registro

    num_resp = validador_registro_respuestas(request, db, registro)

    nombre_dependencia = dependencia_usuario.nombre

    id_jefe_dependencia = dependencia_usuario.id_jefe_dependencia

    usuario_jefe = db(id_jefe_dependencia == auth.user.id).select(db.auth_user.ALL)[0]

    nombre_jefe = usuario_jefe.first_name
    apellido_jefe = usuario_jefe.last_name
    email_jefe = usuario_jefe.email

    nombre_responsable = personal_usuario.f_nombre
    email_responsable = personal_usuario.f_email

    datos_solicitud = [nombre_dependencia, nombre_jefe, apellido_jefe, email_jefe, nombre_responsable, email_responsable, num_resp]
    '''
    #----- AGREGAR RESPUESTA -----#
    if request.post_vars.ci_receptor:
        solic= db(db.t_Solicitud_smydp.id == respuesta.id).select()[0]
        if solic.f_cantidad_conseguida < solic.f_cantidad:
            
            #COMO ESTA SOLICITANDO TIENE UN INGRESO
            #Inventario de quien hace la solicitud calculado por sustancia
            invSolic=db(db.t_Inventario.sustancia==solic.f_sustancia and db.t_Inventario.espacio==solic.f_espacio ).select()
            # Busco el balance de la sustancia a la que me voy a hacer referencia
            balSol=db(db.t_Balance.f_inventario== invSolic[0].id).select()
            #Busco la sustancia a la que estoy sacandole cuentas 
            
        
            total_viejo= invSolic[0].f_existencia
            uso_int_viejo=invSolic[0].f_uso_interno
            espacio = db(db.espacios_fisicos.id == respuesta.f_espacio).select()[0]

            total_nuevo = total_viejo + respuesta.f_cantidad
            uso_interno_nuevo = uso_int_viejo + respuesta.f_cantidad
            
            invSolic[0].update_record(
                    f_existencia=total_nuevo,
                    f_uso_interno=uso_interno_nuevo)
            print(respuesta.f_cantidad)

            db.t_Balance.insert(
                f_cantidad = respuesta.f_cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto='Ingreso',
                f_tipo_ingreso=respuesta.f_calidad[0],
                f_fechaUso=datetime.datetime.now,
                f_medida=invSolic[0].f_medida,
                f_inventario=invSolic[0].id,
                f_sustancia=solic.f_sustancia,
                f_respuesta_solicitud=respuesta.id
            )
        
            #COMO ESTA DANDO TIENE UN CONSUMO
            #Inventario de quien hace la solicitud calculado por sustancia
            invRespo=db(db.t_Inventario.sustancia==solic.f_sustancia and db.t_Inventario.espacio==respuesta.f_espacio).select()
            # Busco el balance de la sustancia a la que me voy a hacer referencia
            balResp=db(db.t_Balance.f_inventario== invRespo[0].id).select()
            #Busco la sustancia a la que estoy sacandole cuentas 
            
        
            total_viejo= invRespo[0].f_existencia
            uso_int_viejo=invRespo[0].f_uso_interno
            espacio = db(db.espacios_fisicos.id == respuesta.f_espacio).select()[0]

            total_nuevo = total_viejo + respuesta.f_cantidad
            uso_interno_nuevo = uso_int_viejo + respuesta.f_cantidad
            
            invRespo[0].update_record(
                    f_existencia=total_nuevo,
                    f_uso_interno=uso_interno_nuevo)
            print(respuesta.f_cantidad)

            db.t_Balance.insert(

                f_cantidad = respuesta.f_cantidad,
                f_cantidad_total=total_nuevo,
                f_concepto='Consumo',
                f_tipo_ingreso=respuesta.f_calidad[0],
                f_fechaUso=datetime.datetime.now,
                f_medida=invRespo[0].f_medida,
                f_inventario=invRespo[0].id,
                f_sustancia=solic.f_sustancia
            )

            ##Cargo la CI del receptor de la sustancia 

            cedula_receptor = int(request.post_vars.ci_receptor)
            inv_id = db(db.t_Respuesta.id == respuesta.id).update(f_responsable_recepcion = cedula_receptor)
            sol[0].update_record(
                        f_cantidad_conseguida=sol[0].f_cantidad_conseguida+respuesta.f_cantidad)

        else :
            if str(respuesta.f_calidad[0])== 'Cesión':
                sol[0].update_record(
                        f_estatus='Completada')
            else :
                sol[0].update_record(
                        f_estatus='Prestamo por devolver')

        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 
    '''

    return dict(respuesta = respuesta,
                medida = medida,
                solicitud = solicitud,
                sustancia = sustancia,
                espacio = espacio,
                datos_solicitud=datos_solicitud,
                responsable = responsable,
                respuestas = respuestas,
                formas = formas,
                esta_autorizado = esta_autorizado,
                espacios = espacios,
                unidades_de_medida=unidades_de_medida
                )

@auth.requires_login(otherwise=URL('modulos', 'login'))
def solicitudes():

    # Lista de sustancias en el catalogo para el modal de agregar sustancia
    # al alcanzar el nivel de espacios fisicos
    sustancias = list(db(db.t_Sustancia.id > 0).select(db.t_Sustancia.ALL))

    # Lista de unidades de medida
    unidades_de_medida = list(db(db.t_Unidad_de_medida.id > 0).select())

    personal_usuario = db(auth.user_id == db.t_Personal.f_usuario).select(db.t_Personal.ALL)[0]

    # Espacios a cargo del usuario actual
    espacios = []
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    if auth.has_membership("TÉCNICO"):
        
        espacios_a_cargo = db(
                (db.t_Personal.f_usuario == auth.user.id) &
                (db.es_encargado.tecnico == db.t_Personal.id) & 
                (db.espacios_fisicos.id == db.es_encargado.espacio_fisico)
                                 ).select()

        espacios = [e.espacios_fisicos for e in espacios_a_cargo]

    else:
        espacios_a_cargo = __get_espacios(user_dep_id)

        for esp in espacios_a_cargo:
                esp_aux = db(
                    (db.espacios_fisicos.id == esp)
                                     ).select()[0]
                espacios.append(esp_aux)
    
            
    
    #----- AGREGAR SOLICITUDES -----#
    if request.post_vars.numRegistro:
        
        
        cantidad = float(request.vars.total)
        unidad = request.vars.unidad
        sustancia = request.vars.sustancia
        uso = request.vars.uso
        justificacion = request.vars.justificacion
        fecha_caducidad = request.vars.fecha_caducidad
        espacio = request.vars.espacio

        fechaCadu=request.vars.fecha_caducidad.split("-")
        fecha_sumi=datetime.datetime(int(fechaCadu[0]),int(fechaCadu[1]),int(fechaCadu[2]))


        fechaHoy = datetime.datetime.now()
        if  fechaHoy > fecha_sumi:
            response.flash = "Fecha de caducidad no puede ser menor a hoy"
            session.flash = response.flash
            redirect(URL('solicitudes'))

        
        numRegistro = request.post_vars.numRegistro
        solicitante = request.post_vars.respSolicitud
        inv_id = db.t_Solicitud_smydp.insert(f_cantidad=cantidad, 
                                            f_responsable_solicitud= personal_usuario.id,
                                            f_cod_registro=numRegistro, 
                                            f_cantidad_conseguida=0,
                                            f_estatus='En espera',
                                            f_justificacion=justificacion,
                                            f_fecha_caducidad=fecha_caducidad,
                                            f_medida=unidad,
                                            f_espacio=espacio,
                                            f_sustancia=sustancia)

        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

    #----- FIN DE AGREGAR SOLICITUDES -----#

    #----- CAMBIO DE ESTADO DE SOLICITUD -----#
    if request.post_vars.idFicha:
        solicitud_a_cambiar = Solicitud(db, auth)
        solicitud_a_cambiar.instanciar(int(request.post_vars.idFicha))
        solicitud_a_cambiar.cambiar_estado(int(request.post_vars.estado), request)
        solicitud_a_cambiar.actualizar(int(request.post_vars.idFicha))

        # if request.post_vars.estado == "1":
        #     solicitud_a_cambiar.fecha_aprobacion = request.now
        #     solicitud_a_cambiar.aprobada_por = auth.user.first_name
        #     solicitud_a_cambiar.actualizar(request.post_vars.idFicha)

        if request.post_vars.estado == "2":
            solicitud_a_cambiar.observaciones = request.post_vars.observaciones
            # solicitud_a_cambiar.elaborada_por = auth.user.first_name
            # solicitud_a_cambiar.fecha_elaboracion = request.now
            solicitud_a_cambiar.actualizar(request.post_vars.idFicha)

            # TODO Quitar la solicitud de la lista de solicitudes luego de que pase a certificarse

            #solicitud_a_cambiar.elaborar_certificacion()

        # if request.post_vars.estado == "-1":
        #     solicitud_a_cambiar.eliminar(int(request.post_vars.idFicha))

        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

    #----- FIN DE CAMBIO DE ESTADO DE SOLICITUD -----#

    #----- ELIMINAR SOLICITUD -----#

    if request.post_vars.eliminar:
        id_a_eliminar = int(request.post_vars.idFicha_eliminar)
        db(id_a_eliminar == db.solicitudes.id).delete()

        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

    #----- FIN DE ELIMINAR SOLICITUD -----#

    #----- DATOS DE SOLICITANTE -----#
    
    dependencia_usuario = db(personal_usuario.f_dependencia == db.dependencias.id).select(db.dependencias.ALL)[0]

    if auth.has_membership(group_id="CLIENTE INTERNO"):
        registro = "FUSB"
    else:
        registro = dependencia_usuario.codigo_registro

    num_registro = validador_registro_solicitudes(request, db, registro)

    nombre_dependencia = dependencia_usuario.nombre

    id_jefe_dependencia = dependencia_usuario.id_jefe_dependencia

    usuario_jefe = db(id_jefe_dependencia == auth.user.id).select(db.auth_user.ALL)[0]

    nombre_jefe = usuario_jefe.first_name
    apellido_jefe = usuario_jefe.last_name
    email_jefe = usuario_jefe.email

    nombre_responsable = personal_usuario.f_nombre
    email_responsable = personal_usuario.f_email

    datos_solicitud = [nombre_dependencia, nombre_jefe, apellido_jefe, email_jefe, nombre_responsable, email_responsable, num_registro]

    #----- GENERACION DE LISTADOS -----#
    listado_de_solicitudes_generadas = ListaSolicitudesHechas(db, datos_solicitud, espacios)

    listado_de_solicitudes_recibidas = ListaSolicitudesRecibidas(db, datos_solicitud, espacios)


    return dict(solicitudes_generadas=listado_de_solicitudes_generadas,
                solicitudes_recibidas=listado_de_solicitudes_recibidas,
                datos_solicitud=datos_solicitud,
                espacios=espacios,
                sustancias=sustancias,
                unidades_de_medida=unidades_de_medida)

@auth.requires_login(otherwise=URL('modulos', 'login'))
def respuestas():

    personal_usuario = db(auth.user_id == db.t_Personal.f_usuario).select(db.t_Personal.ALL)[0]

    # Espacios a cargo del usuario actual
    espacios = []
    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    if auth.has_membership("TÉCNICO"):
        
        espacios_a_cargo = db(
                (db.t_Personal.f_usuario == auth.user.id) &
                (db.es_encargado.tecnico == db.t_Personal.id) & 
                (db.espacios_fisicos.id == db.es_encargado.espacio_fisico)
                                 ).select()

        espacios = [e.espacios_fisicos for e in espacios_a_cargo]

    else:
        espacios_a_cargo = __get_espacios(user_dep_id)

        for esp in espacios_a_cargo:
                esp_aux = db(
                    (db.espacios_fisicos.id == esp)
                                     ).select()[0]
                espacios.append(esp_aux)
    
    #----- CAMBIO DE ESTADO DE SOLICITUD -----#
    if request.post_vars.idFicha:
        solicitud_a_cambiar = Solicitud(db, auth)
        solicitud_a_cambiar.instanciar(int(request.post_vars.idFicha))
        solicitud_a_cambiar.cambiar_estado(int(request.post_vars.estado), request)
        solicitud_a_cambiar.actualizar(int(request.post_vars.idFicha))

        # if request.post_vars.estado == "1":
        #     solicitud_a_cambiar.fecha_aprobacion = request.now
        #     solicitud_a_cambiar.aprobada_por = auth.user.first_name
        #     solicitud_a_cambiar.actualizar(request.post_vars.idFicha)

        if request.post_vars.estado == "2":
            solicitud_a_cambiar.observaciones = request.post_vars.observaciones
            # solicitud_a_cambiar.elaborada_por = auth.user.first_name
            # solicitud_a_cambiar.fecha_elaboracion = request.now
            solicitud_a_cambiar.actualizar(request.post_vars.idFicha)

            # TODO Quitar la solicitud de la lista de solicitudes luego de que pase a certificarse

            #solicitud_a_cambiar.elaborar_certificacion()

        # if request.post_vars.estado == "-1":
        #     solicitud_a_cambiar.eliminar(int(request.post_vars.idFicha))

        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

    #----- FIN DE CAMBIO DE ESTADO DE SOLICITUD -----#

    #----- ELIMINAR SOLICITUD -----#

    if request.post_vars.eliminar:
        id_a_eliminar = int(request.post_vars.idFicha_eliminar)
        db(id_a_eliminar == db.solicitudes.id).delete()

        return redirect(URL(args=request.args, vars=request.get_vars, host=True)) 

    #----- FIN DE ELIMINAR SOLICITUD -----#

    #----- DATOS DE SOLICITANTE -----#
    
    dependencia_usuario = db(personal_usuario.f_dependencia == db.dependencias.id).select(db.dependencias.ALL)[0]

    nombre_dependencia = dependencia_usuario.nombre

    id_jefe_dependencia = dependencia_usuario.id_jefe_dependencia

    usuario_jefe = db(id_jefe_dependencia == auth.user.id).select(db.auth_user.ALL)[0]

    nombre_jefe = usuario_jefe.first_name
    apellido_jefe = usuario_jefe.last_name
    email_jefe = usuario_jefe.email

    nombre_responsable = personal_usuario.f_nombre
    email_responsable = personal_usuario.f_email

    datos_solicitud = [nombre_dependencia, nombre_jefe, apellido_jefe, email_jefe, nombre_responsable, email_responsable]

    #----- GENERACION DE LISTADOS -----#
    respuestas_recibidas = listado_respuestas_recibidas(db, espacios)

    respuestas_enviadas = listado_respuestas_enviadas(db, espacios)


    return dict(respuestas_enviadas=respuestas_enviadas,
                respuestas_recibidas=respuestas_recibidas
                )

@auth.requires_login(otherwise=URL('modulos', 'login'))
def index():
    return locals()

@auth.requires_login(otherwise=URL('modulos', 'login'))
def listado_respuestas_recibidas(db, espacios):
    respuestas = db((db.t_Respuesta.id > 0)).select()
    respuestasRecibidas = {}

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    for resp in respuestas:

        print(resp.f_responsable_recepcion)

        if resp.f_responsable_recepcion == None:

            solicitud = db((db.t_Solicitud_smydp.id == resp.f_solicitud)).select()[0]

            sustancia = db((db.t_Sustancia.id == solicitud.f_sustancia)).select()[0]

            espacio = db((db.espacios_fisicos.id == resp.f_espacio)).select()[0]

            for esp in espacios:
                if espacio.id != esp.id:

                    if not resp.id in respuestasRecibidas:
                        respuestasRecibidas[resp.id] = {
                                            'f_cod_registro': resp.f_cod_registro,
                                            'f_espacio': espacio.codigo,
                                            'f_sustancia': sustancia.f_nombre,
                                            'f_cantidad': resp.f_cantidad,
                                            'f_medida': resp.f_medida,
                                            'f_tipo_respuesta': resp.f_tipo_respuesta,
                                            'f_calidad': resp.f_calidad,
                                            'f_fecha': resp.created_on,
                                            }
    return respuestasRecibidas

@auth.requires_login(otherwise=URL('modulos', 'login'))
def listado_respuestas_enviadas(db, espacios):
    respuestas = db((db.t_Respuesta.id > 0)).select()
    respuestasEnviadas = {}

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    i = 0

    for resp in respuestas:

        solicitud = db((db.t_Solicitud_smydp.id == resp.f_solicitud)).select()[0]

        sustancia = db((db.t_Sustancia.id == solicitud.f_sustancia)).select()[0]

        espacio = db((db.espacios_fisicos.id == resp.f_espacio)).select()[0]

        for esp in espacios:
            if espacio.id == esp.id:

                if not resp.id in respuestasEnviadas:
                    respuestasEnviadas[resp.id] = {
                                        'f_cod_registro': resp.f_cod_registro,
                                        'f_espacio': espacio.codigo,
                                        'f_sustancia': sustancia.f_nombre,
                                        'f_cantidad': resp.f_cantidad,
                                        'f_medida': resp.f_medida,
                                        'f_tipo_respuesta': resp.f_tipo_respuesta,
                                        'f_calidad': resp.f_calidad,
                                        'f_fecha': resp.created_on,
                                        }
    return respuestasEnviadas

@auth.requires_login(otherwise=URL('modulos', 'login'))
def ListaSolicitudesHechas(db, datos, espacios):

    solicitudes = db((db.t_Solicitud_smydp.id > 0)).select()
    solicitudesHechas = {}

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    i = 0

    if auth.has_membership("TÉCNICO"):
        
        for sol in solicitudes:

            sustancia = db((db.t_Sustancia.id == sol.f_sustancia)).select()[0]

            espacio = db(
                            (db.espacios_fisicos.id == sol.f_espacio)
                                 ).select()[0]

            for esp in espacios:
                if espacio.id == esp.id:

                    i += 1
                    solicitudesHechas[sol.id] = {
                                        'f_cod_registro': sol.f_cod_registro,
                                        'f_sustancia': sustancia.f_nombre,
                                        'f_espacio': espacio.codigo,
                                        'f_cantidad': sol.f_cantidad,
                                        'f_fecha': sol.created_on,
                                        'f_estatus':sol.f_estatus
                                        }

    elif auth.has_membership("JEFE DE SECCIÓN"):

        for sol in solicitudes:

            sustancia = db((db.t_Sustancia.id == sol.f_sustancia)).select()[0]
            espacio = db(
                            (db.espacios_fisicos.id == sol.f_espacio)
                                 ).select()[0]

            for esp in espacios:
                if espacio.id == esp.id:

                    i += 1
                    solicitudesHechas[sol.id] = {
                                        'f_cod_registro': sol.f_cod_registro,
                                        'f_sustancia': sustancia.f_nombre,
                                        'f_espacio': espacio.codigo,
                                        'f_cantidad': sol.f_cantidad,
                                        'f_fecha': sol.created_on,
                                        'f_estatus':sol.f_estatus
                                        }

    # Si el usuario no es tecnico, para la base de datos es indiferente su ROL
    # pues la jerarquia de dependencias esta almacenada en la misma tabla
    # con una lista de adyacencias
    else:
        
        for sol in solicitudes:
            sustancia = db((db.t_Sustancia.id == sol.f_sustancia)).select()[0]
            espacio = db(
                            (db.espacios_fisicos.id == sol.f_espacio)
                                 ).select()[0]

            for esp in espacios:
                if espacio.id == esp.id:

                    i += 1
                    solicitudesHechas[sol.id] = {
                                        'f_cod_registro': sol.f_cod_registro,
                                        'f_sustancia': sustancia.f_nombre,
                                        'f_espacio': espacio.codigo,
                                        'f_cantidad': sol.f_cantidad,
                                        'f_fecha': sol.created_on,
                                        'f_estatus': sol.f_estatus
                                        }

    return solicitudesHechas

@auth.requires_login(otherwise=URL('modulos', 'login'))
def ListaSolicitudesRecibidas(db, datos, espacios):

    solicitudes = db((db.t_Solicitud_smydp.id > 0)).select()
    solicitudesRecibidas = {}

    user = db(db.t_Personal.f_usuario == auth.user.id).select()[0]
    user_dep_id = user.f_dependencia

    i = 0

    if auth.has_membership("TÉCNICO"):
        
        for sol in solicitudes:

            if sol.f_estatus[0] != 'Completada':

                sustancia = db((db.t_Sustancia.id == sol.f_sustancia)).select()[0]
                espacio = db(
                                (db.espacios_fisicos.id == sol.f_espacio)
                                     ).select()[0]

                for esp in espacios:
                    if espacio.id != esp.id:

                        for row in db((db.t_Inventario.sustancia == sol.f_sustancia) &
                                      (db.t_Inventario.espacio == esp.id) &
                                      (db.t_Inventario.f_existencia > 0)).select():

                            i += 1
                            solicitudesRecibidas[sol.id] = {
                                                'f_cod_registro': sol.f_cod_registro,
                                                'f_sustancia': sustancia.f_nombre,
                                                'f_espacio': espacio.codigo,
                                                'f_cantidad': sol.f_cantidad,
                                                'f_fecha': sol.created_on,
                                                'f_estatus':sol.f_estatus
                                                }

    elif auth.has_membership("JEFE DE SECCIÓN"):

        for sol in solicitudes:

            if sol.f_estatus[0] != 'Completada':

                sustancia = db((db.t_Sustancia.id == sol.f_sustancia)).select()[0]
                espacio = db(
                                (db.espacios_fisicos.id == sol.f_espacio)
                                     ).select()[0]

                for esp in espacios:
                    if espacio.id != esp.id:

                        for row in db((db.t_Inventario.sustancia == sol.f_sustancia) &
                                      (db.t_Inventario.espacio == esp)).select():

                            i += 1
                            solicitudesRecibidas[sol.id] = {
                                                'f_cod_registro': sol.f_cod_registro,
                                                'f_sustancia': sustancia.f_nombre,
                                                'f_espacio': espacio.codigo,
                                                'f_cantidad': sol.f_cantidad,
                                                'f_fecha': sol.created_on,
                                                'f_estatus':sol.f_estatus
                                                }

    # Si el usuario no es tecnico, para la base de datos es indiferente su ROL
    # pues la jerarquia de dependencias esta almacenada en la misma tabla
    # con una lista de adyacencias
    else:

        for sol in solicitudes:

            if sol.f_estatus[0] != 'Completada':

                sustancia = db((db.t_Sustancia.id == sol.f_sustancia)).select()[0]
                espacio = db(
                                (db.espacios_fisicos.id == sol.f_espacio)
                                     ).select()[0]

                for esp in espacios:
                    if espacio.id != esp.id:

                        for row in db((db.t_Inventario.sustancia == sol.f_sustancia) &
                                      (db.t_Inventario.espacio == esp.id)).select():

                            i += 1
                            solicitudesRecibidas[sol.id] = {
                                                'f_cod_registro': sol.f_cod_registro,
                                                'f_sustancia': sustancia.f_nombre,
                                                'f_espacio': espacio.codigo,
                                                'f_cantidad': sol.f_cantidad,
                                                'f_fecha': sol.created_on,
                                                'f_estatus':sol.f_estatus
                                                }
    return solicitudesRecibidas


def validador_registro_solicitudes(request, db, registro, contador=0):
    anio = str(request.now)[2:4]
    contador = 1 + contador
    digits = (3 - len(str(contador))) * '0' + str(contador)

    registronum = 'SIG-' + registro + "-" + anio + '/' + digits

    check = db(db.t_Solicitud_smydp.f_cod_registro == registronum).count()

    if check != 0:
        return validador_registro_solicitudes(request, db, registro, contador)
    else:
        return registronum

def validador_registro_respuestas(request, db, registro, contador=0):
    anio = str(request.now)[2:4]
    contador = 1 + contador
    digits = (3 - len(str(contador))) * '0' + str(contador)

    registronum = registro + '/' + digits

    check = db(db.t_Respuesta.f_cod_registro == registronum).count()

    if check != 0:
        return validador_registro_respuestas(request, db, registro, contador)
    else:
        return registronum

@auth.requires_login(otherwise=URL('modulos', 'login'))
def sustancias():
    return locals()

# Agrega un nuevo registro a la bitacora de una sustancia desde 
def __agregar_registro_desde_solicitud(registro):

   pass

############################################################################
############################################################################
#       GENERACION DE REPORTES
#############################################################################
############################################################################

 
####################################################################################
##############     GENERACION DE REPORTES   LR7 
####################################################################################
@auth.requires(lambda: __check_role_report())

@auth.requires_login()

def generar_reporte_rl7():

   

    wb = Workbook()
    ws = wb.active
    cen = Alignment(horizontal='center', vertical='distributed')
    rig = Alignment(horizontal='right')
    lef = Alignment(horizontal='left')
    ft1 = Font(name='Arial', size=10, bold=True)
    ft2 = Font(name='Arial', size=10, bold=False)
    ft3 = Font(name='Arial', size=8)
    ws.font = ft2
    now = datetime.datetime.now()

    #mes = '12'
    #year= '2018'
    mes = ((((request.vars.mesR7).replace('(','')).replace(')','')).split(',')[1]).replace(' ','')
    year= request.vars.ayoR7
    #Encabezado
    ws.title = "Informe mensual"
    img = Image("applications/sigulab2/static/images/Logo_ULab.jpg")
    ws.add_image(img, 'A1')
   

    #tamaño de las columnas
    for i in ['A', 'D', 'E','F','G','J','K']:
       ws.column_dimensions[i].width = 10
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 10
    
    #tamaño de las filas
    ws.row_dimensions[13].height = 40
    for i in range(1,13):
        ws.row_dimensions[i].height = 13
    for i in range(14,29):
        ws.row_dimensions[i].height = 12

    #All Merges
    ws.merge_cells(start_row=5,start_column=3,end_row=5,end_column=10)
    ws.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)
    for i in range(13,28):
        ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=4)
        ws.merge_cells(start_row=i,start_column=10,end_row=i,end_column=11)
        
    for i in range(29,33):
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=10)

    #titulos y datos
    z = ['C5', 'J7', 'I9', 'J9', 'K9','B7','B8','B9','B10','B11','I10','J10','K10']
    ws['C5'] = 'INFORME MENSUAL DE SUSTANCIAS QUIMICAS CONTROLADAS'
    ws['J7'] = 'FECHA'
    ws['I9'] = 'DIA'
    ws['J9'] = 'MES'
    ws['K9'] = 'AÑO'



    for i in range(5):
        ws[z[i]].font = ft1
        ws[z[i]].alignment = cen

    ws['B7'] = 'OPERADOR:'
    ws['B8'] = 'LICENCIA:'
    ws['B9'] = 'PERMISO DEL CICPC:'
    ws['B10'] = 'RIF:'
    ws['B11'] = 'MES-AÑO:'
   

    for i in range(5,10):
        ws[z[i]].font = ft1
        ws[z[i]].alignment = rig
        

    ws['I10'] = now.day
    ws['J10'] = now.month
    ws['K10'] = now.year
    

    for i in range(10,13):
        ws[z[i]].font = ft2
        ws[z[i]].alignment = cen
       

    ws['A28'] = 'Nota:'
    ws['A28'].font = ft1
    ws['A28'].alignment = lef
    
    #mes = request.vars['m']
    #year= request.vars['y']
    
    w = ['C7', 'C8', 'C9', 'C10', 'C11','A13','B13','D13','E13','F13','G13','H13','I13','J13']

    for i in range(5):
        ws[w[i]].font = ft2

    ws['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
    ws['C8'] = '2014LIC0256'
    ws['C9'] = 'No. 1311'
    ws['C10'] = 'G-20000063-5'
    ws['C11'] = mes+'/'+year
    ws['A13'] = 'N°'

    ws['B13'] = 'Sustancia Química Controlada'

  

    ws['E13'] = 'Saldo Físico Inicial'

    ws['F13'] = 'Total Entradas'

    ws['G13'] = 'Total Salidas'

    ws['H13'] = 'Saldo Físico Final'

    ws['I13'] = 'Unidad de Medida'

    ws['J13'] = 'Observaciones'




    for i in range(5,14):
        ws[w[i]].font = ft1
        ws[w[i]].alignment = cen
        


    x = ['A14','A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26']
    y = ['01','02','03','04','05','06','07','08','09','10','11','12','13']
    for i in range(0,13):
        ws[x[i]] = y[i]
        ws[x[i]].font = ft3
        ws[x[i]].alignment = cen
       
    x = ['B14','B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26']
    
    ####################################################################
    ######### FIN DEL ENCABEZADO
    ####################################################################
    
    # CONSULTA DE LAS SUSTANCIAS REGULADAS LR4 Y QUE SE LES HA APERTURADO BALANCE 
    # EN EL SISTAMA 
    sustContl7= db((db.t_Sustancia.f_control=="RL7")or (db.t_Sustancia.f_control=="RL4 y RL7")).select()
    sustBit=db((db.t_Balance.f_fechaUso.year()==int(year))&(db.t_Balance.f_fechaUso.month()==int(mes))).select(orderby=~db.t_Balance.f_fechaUso)
    medidas={}
    ids={}
    entradas = {}
    salidas = {}
    totalIni={}
    totalFin={}
    
    for suCo in sustContl7:
        aux=False
        auxEnt=0
        auxSal=0
        auxFecIn=31
        auxFecFi=1 

        for suFe in sustBit:
            if (suCo.id== suFe['f_sustancia'] and not(aux)):
                ids[str(suCo.id)]=suCo.f_nombre
                medidas[str(suCo.id)]= suFe['f_medida']
                aux=True
                salidas[str(suCo.id)]=0
                entradas[str(suCo.id)]= 0
            if (suCo.id== suFe['f_sustancia']): 
                if ( suFe['f_concepto']==['Ingreso']):
                    auxEnt+= float(suFe['f_cantidad'])
                    entradas[str(suCo.id)]= auxEnt
                elif ( suFe['f_concepto']==['Consumo']):
                    auxSal+=float(suFe['f_cantidad'])
                    salidas[str(suCo.id)]= auxSal 

                    
                if (int(str(suFe['f_fechaUso']).split('-')[2])<=auxFecIn):
                    totalIni[str(suCo.id)]=float(suFe['f_cantidad_total'])
                    auxFecIn=int(str(suFe['f_fechaUso']).split('-')[2] )

                if (int(str(suFe['f_fechaUso']).split('-')[2])>=auxFecFi):
                    totalFin[str(suCo.id)]=float(suFe['f_cantidad_total'])
                    auxFecFi=int(str(suFe['f_fechaUso']).split('-')[2] )        
                   # print(str(suFe['f_cantidad_total']))
    
    ## CALCULANDO LA CANTIDAD DE TRANSACCIONES SE REALIZARON DE LA SUSTANCIA
   
   
    
    y=0;
    # CARGANDO LOS NOMBRES AL EXCEL
    for i,names in ids.items():
        if y<13:
            ws[x[y]] = names
            ws[x[y]].font = ft3
            y=y+1
            
    x = ['E14','E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26']

    ##
   ## SALDO FISICO INICIAL

    y=0;
    for i,names in totalIni.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1 
            else:
                ws[x[y]] = float(names)
                ws[x[y]].font = ft3
                y=y+1 

         



    ###
    ### TOTAL DE ENTRADAS

    x = ['F14','F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26']
    y=0;
    for i,names in entradas.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1 
            else:
                ws[x[y]] = float(names)
                ws[x[y]].font = ft3
                y=y+1 

            
        
    ###
    ### TOTAL DE SALIDAS

    f = ['G14','G15','G16','G17','G18','G19','G20','G21','G22','G23','G24','G25','G26']
    y=0;
    for i,names in salidas.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[f[y]] = auxP
                ws[f[y]].font = ft3
                y=y+1 
            else:
                ws[f[y]] = float(names)
                ws[f[y]].font = ft3
                y=y+1 
         

    ###
    ### TOTAL DE SALIDAS

    x = ['H14','H15','H16','H17','H18','H19','H20','H21','H22','H23','H24','H25','H26']

    y=0;
    for i,names in totalFin.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1 
            else:
                ws[x[y]] = float(names)
                ws[x[y]].font = ft3
                y=y+1 
    

    #####################################
    # RELLENANDO LA UNIDAD DE MEDIDA 
    #####################################

    x = ['I14','I15','I16','I17','I18','I19','I20','I21','I22','I23','I24','I25','I26']

    y=0;
    for i,medi in medidas.items():
        query=db((db.t_Unidad_de_medida.id==int(medi))).select(db.t_Unidad_de_medida.f_abreviatura)
        if y<13:
            if str(query[0].f_abreviatura)=="ml":
                ws[x[y]] = "l"
            elif str(query[0].f_abreviatura)=="g":
                ws[x[y]] = "kg"
            else:
                ws[x[y]] = str((query[0].f_abreviatura))
            ws[x[y]].font = ft3

            y=y+1

    
       

    #Pie de Pagina
    ws['A29'] = '1. Los saldos serán reportados en:'
    ws['A30'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
    ws['A31'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
    ws['A32'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'
  


    ###########################################################################
    ###########################################################################
    #           REPORTES INDIVIDUALES 
    ##########################################################################
    ##########################################################################
    namesList = {}
    bitacora=[]
    for suCo in sustContl7:
        suAux= db((db.t_Balance.f_sustancia==suCo.id)).select()
        aux=0;
        nameBol=False
        for j in suAux:
            
            if(j.f_fechaUso.month==int(mes) and j.f_fechaUso.year==int(year)):
                aux=aux+1
        for i,n in ids.items():
            if (suCo.f_nombre==n):
                namesList[suCo.id]= aux 
        bitacora.append(aux)  

    contador=0
        
    for neId,n in ids.items():
        while ( len(n)>=31):
            h=n.split(' ')
            h.pop()
            n=' '.join(map(str,h))
        try:
            n=unicode(n,"utf-8")
        except:
            pass
        ws2 = wb.create_sheet(n)
        # Encabezado 

        ws2.title = n
        img = Image("applications/sigulab2/static/images/Logo_ULab.jpg")
        ws2.add_image(img, 'A1')

        #tamaño de las columnas
        for i in ['A', 'D', 'K','G','H','I']:
            ws2.column_dimensions[i].width = 9
        ws2.column_dimensions['B'].width = 9
        ws2.column_dimensions['C'].width = 17.5
        ws2.column_dimensions['E'].width = 17.5
        ws2.column_dimensions['F'].width = 17.5
        ws2.column_dimensions['J'].width = 17.5


        #tamaño de las filas
        ws2.row_dimensions[14].height = 40
        for i in range(1,14):
            ws2.row_dimensions[i].height = 13
        for i in range(15,42):
            ws2.row_dimensions[i].height = 13

  

        #All Merges
        ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
        ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

        #titulos y datos
        z = ['B5', 'G7', 'F8', 'G8', 'H8','B7','B8','B9','B10','B11','B12','F9','G9','H9']
        ws2['B5'] = 'INFORME DE REPORTE DIARIO DE SUSTANCIAS QUIMICAS CONTROLADAS'
        ws2['G7'] = 'FECHA'
        ws2['F8'] = 'DIA'
        ws2['G8'] = 'MES'
        ws2['H8'] = 'AÑO'


        for i in range(5):
            ws2[z[i]].font = ft1
            ws2[z[i]].alignment = cen


        ws2['B7'] = 'OPERADOR:'
        ws2['B8'] = 'LICENCIA:'
        ws2['B9'] = 'RIF:'
        ws2['B10'] = 'SUSTANCIA:'
        ws2['B11'] = 'UNIDAD DE MEDIDA:'
        ws2['B12'] = 'MES-AÑO:'



        for i in range(5,11):
            ws2[z[i]].font = ft1
            ws2[z[i]].alignment = rig

        ws2['F9'] = now.day
        ws2['G9'] = now.month
        ws2['H9'] = now.year


        for i in range(11,14):
            ws2[z[i]].font = ft2
            ws2[z[i]].alignment = cen


        ws2['A36'] = 'Nota:'
        ws2['A36'].font = ft1
        ws2['A36'].alignment = lef


        w = ['C7', 'C8', 'C9', 'C10', 'C11','C12','A14','B14','C14','D14','E14','F14','G14','H14','I14','J14',]

        for i in range(6):
            ws2[w[i]].font = ft2

        ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
        ws2['C8'] = '2014LIC0256'
        ws2['C9'] = 'G-20000063-5'
        ws2['C10'] =  n.upper()
        ws2['C11'] = 'med' #aqui va la unidad de medida 
        ws2['C12'] = mes+'-'+year 
        #query=db((db.t_Unidad_de_medida.id==medidas[str(i)])).select(db.t_Unidad_de_medida.f_abreviatura)
        ws2['C11'] =' '



        ws2['A14'] = 'Asiento'

        ws2['B14'] = 'Fecha'

        ws2['C14'] = 'Documento Nro'

        ws2['D14'] = 'RIF o Cédula de identidad'

        ws2['E14'] = 'Nombre de la persona natural o juridica '

        ws2['F14'] = 'Descripción (de acuerdo a su actividad)'

        ws2['G14'] = 'Entrada'

        ws2['H14'] = 'Salida'

        ws2['I14'] = 'Saldo'

        ws2['J14'] = 'Observaciones'

        for i in range(5,16):
            ws2[w[i]].font = ft1
            ws2[w[i]].alignment = cen

        x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34']
        y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20']
        for i in range(0,20):
            ws2[x[i]] = y[i]
            ws2[x[i]].font = ft2
            ws2[x[i]].alignment = cen
        
        x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34']
        
        fechasImdiv={}
        sufeAux=[]
        consumoIndiv={}
        consAux=[]
        ingresoIndiv={}
        ingAux=[]
        totalInd={}
        auxTotal=[]
        auxTotalNu=0
        for suFe in sustBit:
            if (int(neId)== int(suFe['f_sustancia'])): 
                sufeAux.append(suFe['f_fechaUso'])
                medida = suFe['f_medida']
                if ( suFe['f_concepto']==['Ingreso']):
                    ingAux.append(float(suFe['f_cantidad']))
                    auxTotalNu+=float(suFe['f_cantidad'])
                    auxTotal.append(float(auxTotalNu))

                elif ( suFe['f_concepto']==['Consumo']):
                    consAux.append(float(suFe['f_cantidad']))
                    auxTotalNu-=float(suFe['f_cantidad'])
                    auxTotal.append(float(auxTotalNu))

        
        fechasImdiv[str(neId)] = sufeAux
        ingresoIndiv[str(neId)]= ingAux
        consumoIndiv[str(neId)]= consAux 
        totalInd[str(neId)]=auxTotal
        h=0
        for i in fechasImdiv[str(neId)]:
            ws2[x[h]] = str(i)
            ws2[x[h]].font = ft2
            ws2[x[h]].alignment = cen
            h+=1


        y = ['G15','G16','G17','G18','G19','G20','G21','G22','G23','G24','G25','G26','G27']
        h1=0
        for i in ingresoIndiv[str(neId)]:
            query=db(db.t_Unidad_de_medida.id==int(medida)).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws2[y[h1]] = float(i)/1000
                ws2[y[h1]].font = ft2
                ws2[y[h1]].alignment = cen
                h1+=1
            else:
                ws2[y[h1]] = str(i)
                ws2[y[h1]].font = ft2
                ws2[y[h1]].alignment = cen
                h1+=1

           
        z = ['H15','H16','H17','H18','H19','H20','H21','H22','H23','H24','H25','H26','H27']
        h2=0
        for i in consumoIndiv[str(neId)]:
            query=db(db.t_Unidad_de_medida.id==int(medida)).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws2[z[h2]] = float(i)/1000
                ws2[z[h2]].font = ft2
                ws2[z[h2]].alignment = cen
                h2+=1

            else:
                ws2[z[h2]] = str(i)
                ws2[z[h2]].font = ft2
                ws2[z[h2]].alignment = cen
                h2+=1


        w = ['I15','I16','I17','I18','I19','I20','I21','I22','I23','I24','I25','I26','I27']
        h3=0
        for i in totalInd[str(neId)]:
            query=db(db.t_Unidad_de_medida.id==int(medida)).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws2[w[h3]] = float(i)/1000
                ws2[w[h3]].font = ft2
                ws2[w[h3]].alignment = cen
                h3+=1

            else:
                ws2[w[h3]] = str(i)
                ws2[w[h3]].font = ft2
                ws2[w[h3]].alignment = cen
                h3+=1

           


    #Pie de Pagina
        ws2['A37'] = '1. Los saldos serán reportados en:'
        ws2['A38'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
        ws2['A39'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
        ws2['A40'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'








 
    wb.save('Reporte Universidad Simon Bolivar_l7.xlsx')
    response.stream('Reporte Universidad Simon Bolivar_l7.xlsx',attachment=True, filename='Reporte Universidad Simon Bolivar.xlsx')
    return locals()


  
####################################################################################
##############     GENERACION DE REPORTES   LR4
####################################################################################
@auth.requires(lambda: __check_role_report())

@auth.requires_login()

def generar_reporte_rl4():

    
    wb = Workbook()
    ws = wb.active
    cen = Alignment(horizontal='center', vertical='distributed')
    rig = Alignment(horizontal='right')
    lef = Alignment(horizontal='left')
    ft1 = Font(name='Arial', size=10, bold=True)
    ft2 = Font(name='Arial', size=10, bold=False)
    ft3 = Font(name='Arial', size=8)
    ws.font = ft2
    now = datetime.datetime.now()

    #mes = '12'
    #year= '2018'
    mes = ((((request.vars.mesR4).replace('(','')).replace(')','')).split(',')[1]).replace(' ','')
    year= request.vars.ayoR4
    #Encabezado
    ws.title = "Informe mensual"
    img = Image("gob.jpg")
    ws.add_image(img, 'A1')
   

    #tamaño de las columnas
    for i in ['A', 'D', 'E','F','G','J','K']:
       ws.column_dimensions[i].width = 10
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 10
    
    #tamaño de las filas
    ws.row_dimensions[13].height = 40
    for i in range(1,13):
        ws.row_dimensions[i].height = 13
    for i in range(14,29):
        ws.row_dimensions[i].height = 12

    #All Merges
    ws.merge_cells(start_row=5,start_column=3,end_row=5,end_column=10)
    ws.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)
    for i in range(13,28):
        ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=4)
        ws.merge_cells(start_row=i,start_column=10,end_row=i,end_column=11)
        
    for i in range(29,33):
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=10)

    #titulos y datos
    z = ['C5', 'J7', 'I9', 'J9', 'K9','B7','B8','B9','B10','B11','I10','J10','K10']
    ws['C5'] = 'INFORME MENSUAL DE SUSTANCIAS QUIMICAS CONTROLADAS'
    ws['J7'] = 'FECHA'
    ws['I9'] = 'DIA'
    ws['J9'] = 'MES'
    ws['K9'] = 'AÑO'



    for i in range(5):
        ws[z[i]].font = ft1
        ws[z[i]].alignment = cen

    ws['B7'] = 'OPERADOR:'
    ws['B8'] = 'LICENCIA:'
    ws['B9'] = 'PERMISO DEL CICPC:'
    ws['B10'] = 'RIF:'
    ws['B11'] = 'MES-AÑO:'
   

    for i in range(5,10):
        ws[z[i]].font = ft1
        ws[z[i]].alignment = rig
        

    ws['I10'] = now.day
    ws['J10'] = now.month
    ws['K10'] = now.year
    

    for i in range(10,13):
        ws[z[i]].font = ft2
        ws[z[i]].alignment = cen
       

    ws['A28'] = 'Nota:'
    ws['A28'].font = ft1
    ws['A28'].alignment = lef
    
    #mes = request.vars['m']
    #year= request.vars['y']
    
    w = ['C7', 'C8', 'C9', 'C10', 'C11','A13','B13','D13','E13','F13','G13','H13','I13','J13']

    for i in range(5):
        ws[w[i]].font = ft2

    ws['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
    ws['C8'] = '2014LIC0256'
    ws['C9'] = 'No. 1311'
    ws['C10'] = 'G-20000063-5'
    ws['C11'] = mes+'/'+year
    ws['A13'] = 'N°'

    ws['B13'] = 'Sustancia Química Controlada'

  

    ws['E13'] = 'Saldo Físico Inicial'

    ws['F13'] = 'Total Entradas'

    ws['G13'] = 'Total Salidas'

    ws['H13'] = 'Saldo Físico Final'

    ws['I13'] = 'Unidad de Medida'

    ws['J13'] = 'Observaciones'




    for i in range(5,14):
        ws[w[i]].font = ft1
        ws[w[i]].alignment = cen
        


    x = ['A14','A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26']
    y = ['01','02','03','04','05','06','07','08','09','10','11','12','13']
    for i in range(0,13):
        ws[x[i]] = y[i]
        ws[x[i]].font = ft3
        ws[x[i]].alignment = cen
       
    x = ['B14','B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26']
    
    ####################################################################
    ######### FIN DEL ENCABEZADO
    ####################################################################
    
    # CONSULTA DE LAS SUSTANCIAS REGULADAS LR4 Y QUE SE LES HA APERTURADO BALANCE 
    # EN EL SISTAMA 
    sustContl7= db((db.t_Sustancia.f_control=="RL4")or (db.t_Sustancia.f_control=="RL4 y RL7")).select()
    sustBit=db((db.t_Balance.f_fechaUso.year()==int(year))&(db.t_Balance.f_fechaUso.month()==int(mes))).select(orderby=~db.t_Balance.f_fechaUso)
    medidas={}
    ids={}
    entradas = {}
    salidas = {}
    totalIni={}
    totalFin={}
    
    for suCo in sustContl7:
        aux=False
        auxEnt=0
        auxSal=0
        auxFecIn=31
        auxFecFi=1 

        for suFe in sustBit:
            if (suCo.id== suFe['f_sustancia'] and not(aux)):
                ids[str(suCo.id)]=suCo.f_nombre
                medidas[str(suCo.id)]= suFe['f_medida']
                aux=True
                salidas[str(suCo.id)]=0
                entradas[str(suCo.id)]= 0
            if (suCo.id== suFe['f_sustancia']): 
                if ( suFe['f_concepto']==['Ingreso']):
                    auxEnt+= float(suFe['f_cantidad'])
                    entradas[str(suCo.id)]= auxEnt
                elif ( suFe['f_concepto']==['Consumo']):
                    auxSal+=float(suFe['f_cantidad'])
                    salidas[str(suCo.id)]= auxSal  
                if (int(str(suFe['f_fechaUso']).split('-')[2])<=auxFecIn):
                    totalIni[str(suCo.id)]=float(suFe['f_cantidad_total'])
                    auxFecIn=int(str(suFe['f_fechaUso']).split('-')[2] )

                if (int(str(suFe['f_fechaUso']).split('-')[2])>=auxFecFi):
                    totalFin[str(suCo.id)]=float(suFe['f_cantidad_total'])
                    auxFecFi=int(str(suFe['f_fechaUso']).split('-')[2] )        
                   # print(str(suFe['f_cantidad_total']))
    
    ## CALCULANDO LA CANTIDAD DE TRANSACCIONES SE REALIZARON DE LA SUSTANCIA
   
   
    
    y=0;
    # CARGANDO LOS NOMBRES AL EXCEL
    for i,names in ids.items():
        if y<13:
            ws[x[y]] = names
            ws[x[y]].font = ft3
            y=y+1
            
    x = ['E14','E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26']

    ##
   ## SALDO FISICO INICIAL

    y=0;
    for i,names in totalIni.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1
            else:
                ws[x[y]].font = ft3
                auxP=float(names) 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1    


    ###
    ### TOTAL DE ENTRADAS

    x = ['F14','F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26']
    y=0;
    for i,names in entradas.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1
            else:
                ws[x[y]].font = ft3
                auxP=float(names) 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1    
            
        
    ###
    ### TOTAL DE SALIDAS

    f = ['G14','G15','G16','G17','G18','G19','G20','G21','G22','G23','G24','G25','G26']
    y=0;
    for i,names in salidas.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[f[y]].font = ft3
                auxP=float(names)/1000 
                ws[f[y]] = auxP
                ws[f[y]].font = ft3
                y=y+1
            else:
                ws[f[y]].font = ft3
                auxP=float(names) 
                ws[f[y]] = auxP
                ws[f[y]].font = ft3
                y=y+1    
           

    ###
    ### TOTAL DE SALIDAS

    x = ['H14','H15','H16','H17','H18','H19','H20','H21','H22','H23','H24','H25','H26']

    y=0;
    for i,names in totalFin.items():
        if y<13:
            query=db((db.t_Unidad_de_medida.id==int(medidas[str(i)]))).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws[x[y]].font = ft3
                auxP=float(names)/1000 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1
            else:
                ws[x[y]].font = ft3
                auxP=float(names) 
                ws[x[y]] = auxP
                ws[x[y]].font = ft3
                y=y+1    
    

    #####################################
    # RELLENANDO LA UNIDAD DE MEDIDA 
    #####################################

    x = ['I14','I15','I16','I17','I18','I19','I20','I21','I22','I23','I24','I25','I26']

    y=0;
    for i,medi in medidas.items():
        query=db((db.t_Unidad_de_medida.id==int(medi))).select(db.t_Unidad_de_medida.f_abreviatura)
        if y<13:
            if str(query[0].f_abreviatura)=='ml':
                ws[x[y]] = "l"
            elif str(query[0].f_abreviatura)=='g':
                ws[x[y]] = "kg"
            else:
                ws[x[y]] = str((query[0].f_abreviatura))
            ws[x[y]].font = ft3

            y=y+1

    
       

    #Pie de Pagina
    ws['A29'] = '1. Los saldos serán reportados en:'
    ws['A30'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
    ws['A31'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
    ws['A32'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'
  


    ###########################################################################
    ###########################################################################
    #           REPORTES INDIVIDUALES 
    ##########################################################################
    ##########################################################################
    namesList = {}
    bitacora=[]
    for suCo in sustContl7:
        suAux= db((db.t_Balance.f_sustancia==suCo.id)).select()
        aux=0;
        nameBol=False
        for j in suAux:
            
            if(j.f_fechaUso.month==int(mes) and j.f_fechaUso.year==int(year)):
                aux=aux+1
        for i,n in ids.items():
            if (suCo.f_nombre==n):
                namesList[suCo.id]= aux 
        bitacora.append(aux)  

    contador=0
        
    for neId,n in ids.items():

        while ( len(n)>=31):
            h=n.split(' ')
            h.pop()
            n=' '.join(map(str,h))
        try:
            n=unicode(n,"utf-8")
        except:
            pass
        ws2 = wb.create_sheet(n)
        # Encabezado 

        ws2.title = n
        
        img = Image("gob.jpg")
        ws2.add_image(img, 'A1')

        #tamaño de las columnas
        for i in ['A', 'D', 'K','G','H','I']:
            ws2.column_dimensions[i].width = 9
        ws2.column_dimensions['B'].width = 9
        ws2.column_dimensions['C'].width = 17.5
        ws2.column_dimensions['E'].width = 17.5
        ws2.column_dimensions['F'].width = 17.5
        ws2.column_dimensions['J'].width = 17.5


        #tamaño de las filas
        ws2.row_dimensions[14].height = 40
        for i in range(1,14):
            ws2.row_dimensions[i].height = 13
        for i in range(15,42):
            ws2.row_dimensions[i].height = 13

  

        #All Merges
        ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
        ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

        #titulos y datos
        z = ['B5', 'G7', 'F8', 'G8', 'H8','B7','B8','B9','B10','B11','B12','F9','G9','H9']
        ws2['B5'] = 'INFORME DE REPORTE DIARIO DE SUSTANCIAS QUIMICAS CONTROLADAS'
        ws2['G7'] = 'FECHA'
        ws2['F8'] = 'DIA'
        ws2['G8'] = 'MES'
        ws2['H8'] = 'AÑO'


        for i in range(5):
            ws2[z[i]].font = ft1
            ws2[z[i]].alignment = cen


        ws2['B7'] = 'OPERADOR:'
        ws2['B8'] = 'LICENCIA:'
        ws2['B9'] = 'RIF:'
        ws2['B10'] = 'SUSTANCIA:'
        ws2['B11'] = 'UNIDAD DE MEDIDA:'
        ws2['B12'] = 'MES-AÑO:'



        for i in range(5,11):
            ws2[z[i]].font = ft1
            ws2[z[i]].alignment = rig

        ws2['F9'] = now.day
        ws2['G9'] = now.month
        ws2['H9'] = now.year


        for i in range(11,14):
            ws2[z[i]].font = ft2
            ws2[z[i]].alignment = cen


        ws2['A36'] = 'Nota:'
        ws2['A36'].font = ft1
        ws2['A36'].alignment = lef


        w = ['C7', 'C8', 'C9', 'C10', 'C11','C12','A14','B14','C14','D14','E14','F14','G14','H14','I14','J14',]

        for i in range(6):
            ws2[w[i]].font = ft2

        ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
        ws2['C8'] = '2014LIC0256'
        ws2['C9'] = 'G-20000063-5'
        ws2['C10'] =  n.upper()
        ws2['C11'] = 'med' #aqui va la unidad de medida 
        ws2['C12'] = mes+'-'+year 
        #query=db((db.t_Unidad_de_medida.id==medidas[str(i)])).select(db.t_Unidad_de_medida.f_abreviatura)
        ws2['C11'] =' '



        ws2['A14'] = 'Asiento'

        ws2['B14'] = 'Fecha'

        ws2['C14'] = 'Documento Nro'

        ws2['D14'] = 'RIF o Cédula de identidad'

        ws2['E14'] = 'Nombre de la persona natural o juridica '

        ws2['F14'] = 'Descripción (de acuerdo a su actividad)'

        ws2['G14'] = 'Entrada'

        ws2['H14'] = 'Salida'

        ws2['I14'] = 'Saldo'

        ws2['J14'] = 'Observaciones'

        for i in range(5,16):
            ws2[w[i]].font = ft1
            ws2[w[i]].alignment = cen

        x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34']
        y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20']
        for i in range(0,20):
            ws2[x[i]] = y[i]
            ws2[x[i]].font = ft2
            ws2[x[i]].alignment = cen
        
        x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34']
        
        fechasImdiv={}
        sufeAux=[]
        consumoIndiv={}
        consAux=[]
        ingresoIndiv={}
        ingAux=[]
        totalInd={}
        auxTotal=[]
        auxTotalNu=0
        
        
        for suFe in sustBit:
            if (int(neId)== int(suFe['f_sustancia'])): 
                sufeAux.append(suFe['f_fechaUso'])
                medida = suFe['f_medida']
                if ( suFe['f_concepto']==['Ingreso']):
                    ingAux.append(float(suFe['f_cantidad']))
                    consAux.append(0)
                    auxTotalNu+=float(suFe['f_cantidad'])
                    auxTotal.append(float(auxTotalNu))
                elif ( suFe['f_concepto']==['Consumo']):
                    consAux.append(float(suFe['f_cantidad']))
                    ingAux.append(0)
                    auxTotalNu-=float(suFe['f_cantidad'])
                    auxTotal.append(float(auxTotalNu))
        

                   
        fechasImdiv[str(neId)] = sufeAux
        ingresoIndiv[str(neId)]= ingAux
        consumoIndiv[str(neId)]= consAux 
        totalInd[str(neId)]=auxTotal

        h=0
        for i in fechasImdiv[str(neId)]:
            ws2[x[h]] = str(i)
            ws2[x[h]].font = ft2
            ws2[x[h]].alignment = cen
            h+=1


        y = ['G15','G16','G17','G18','G19','G20','G21','G22','G23','G24','G25','G26','G27']
        h1=0
        for i in ingresoIndiv[str(neId)]:
            query=db(db.t_Unidad_de_medida.id==int(medida)).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws2[y[h1]] = float(i)/1000
                ws2[y[h1]].font = ft2
                ws2[y[h1]].alignment = cen
                h1+=1
            else:
                ws2[y[h1]] = str(i)
                ws2[y[h1]].font = ft2
                ws2[y[h1]].alignment = cen
                h1+=1

            

        z = ['H15','H16','H17','H18','H19','H20','H21','H22','H23','H24','H25','H26','H27']
        h2=0
        for i in consumoIndiv[str(neId)]:
            query=db(db.t_Unidad_de_medida.id==int(medida)).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws2[z[h2]] = float(i)/1000
                ws2[z[h2]].font = ft2
                ws2[z[h2]].alignment = cen
                h2+=1
            else:
                ws2[z[h2]] = str(i)
                ws2[z[h2]].font = ft2
                ws2[z[h2]].alignment = cen
                h2+=1
            

        w = ['I15','I16','I17','I18','I19','I20','I21','I22','I23','I24','I25','I26','I27']
        h3=0
        for i in totalInd[str(neId)]:

            query=db(db.t_Unidad_de_medida.id==int(medida)).select(db.t_Unidad_de_medida.f_abreviatura)
            if ((str(query[0].f_abreviatura)=='ml') or (str(query[0].f_abreviatura)=='g')):
                ws2[w[h3]] = float(i)/1000
                ws2[w[h3]].font = ft2
                ws2[w[h3]].alignment = cen
                h3+=1
            else:
                ws2[w[h3]] = str(i)
                ws2[w[h3]].font = ft2
                ws2[w[h3]].alignment = cen
                h3+=1
            


    #Pie de Pagina
        ws2['A37'] = '1. Los saldos serán reportados en:'
        ws2['A38'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
        ws2['A39'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
        ws2['A40'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'








 
    wb.save('Reporte Universidad Simon Bolivar_l7.xlsx')
    response.stream('Reporte Universidad Simon Bolivar_l7.xlsx',attachment=True, filename='Reporte Universidad Simon Bolivar.xlsx')
    return locals()